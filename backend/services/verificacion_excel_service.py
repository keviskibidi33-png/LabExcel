"""
Servicio para generar archivos Excel de verificación de muestras cilíndricas
Basado en el template VERIFICACION CONCRETO - AUTOMATIZADO.xlsx
"""

import os
import shutil
import logging
from datetime import datetime
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models import VerificacionMuestras, MuestraVerificada

logger = logging.getLogger(__name__)

class VerificacionExcelService:
    """
    Servicio para generar archivos Excel de verificación de muestras cilíndricas.
    
    Utiliza un template predefinido y llena los datos específicos de cada verificación
    respetando el formato y estructura del documento original.
    """
    
    # Mapeo de columnas para el template V03
    # Orden exacto según especificación del usuario (sin celdas fusionadas):
    # PLANITUD: C.SUPERIOR (M), C.INFERIOR (N), DEPRESIONES (O)
    # ACCION A REALIZAR (P), CONFORMIDAD (Q)
    # LONGITUD 1 (R), LONGITUD 2 (S), LONGITUD 3 (T)
    # MASA MUESTRA AIRE (U), PESAR/NO PESAR (V)
    COLUMNS = {
        'numero': 1,                    # A - N°
        'codigo_lem': 2,                # B - Código LEM
        'tipo_testigo': 3,              # C - TIPO DE TESTIGO
        'diametro_1': 4,                # D - Diámetro 1 (mm)
        'diametro_2': 5,                # E - Diámetro 2 (mm)
        'tolerancia_porcentaje': 6,     # F - ΔΦ 2%>
        'aceptacion_diametro': 7,       # G - Aceptación
        # PERPENDICULARIDAD (5 columnas)
        'perpendicularidad_sup1': 8,    # H - SUP 1 Aceptacion
        'perpendicularidad_sup2': 9,    # I - SUP 2 Aceptacion
        'perpendicularidad_inf1': 10,   # J - INF 1 Aceptacion
        'perpendicularidad_inf2': 11,   # K - INF 2 Aceptacion
        'perpendicularidad_medida': 12, # L - MEDIDA < 0.5°
        # PLANITUD (sin fusionar)
        'planitud_superior': 13,        # M - C. SUPERIOR < 0.05 mm Aceptacion
        'planitud_inferior': 14,        # N - C. INFERIOR < 0.05 mm Aceptacion
        'planitud_depresiones': 15,     # O - Depresiones ≤ 5 mm Aceptacion
        'accion': 16,                   # P - ACCIÓN A REALIZAR
        'conformidad': 17,              # Q - Conformidad
        # LONGITUD (3 columnas)
        'longitud_1': 18,               # R - Longitud 1 (mm)
        'longitud_2': 19,               # S - Longitud 2 (mm)
        'longitud_3': 20,               # T - Longitud 3 (mm)
        'masa': 21,                     # U - Masa muestra aire (g) - dígitos
        'pesar': 22                    # V - Pesar / No pesar
    }
    
    def __init__(self):
        """Inicializa el servicio con las rutas de template y salida."""
        # Nuevo template V03 - Archivo xlsx en la raíz de templates
        self.template_path = "templates/VERIFICACION CONCRETO - AUTOMATIZADO.xlsx"
        self.output_dir = "backend/output"
        
        # Asegurar que el directorio de salida existe
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generar_excel_verificacion(self, verificacion: VerificacionMuestras) -> str:
        """
        Genera un archivo Excel para la verificación de muestras cilíndricas.
        
        Args:
            verificacion: Objeto VerificacionMuestras con los datos a llenar
            
        Returns:
            str: Ruta del archivo Excel generado
            
        Raises:
            ValueError: Si hay error en la generación del archivo
        """
        try:
            # Crear archivo de salida con timestamp único
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"verificacion_muestras_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            # Copiar el template para preservar formato
            shutil.copy2(self.template_path, filepath)
            
            # Cargar el archivo copiado
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Llenar datos específicos respetando el formato del template
            self._llenar_datos_template(ws, verificacion)
            
            # Guardar el archivo
            wb.save(filepath)
            
            # Actualizar la ruta en la base de datos
            verificacion.archivo_excel = filepath
            
            logger.info(f"Archivo Excel generado exitosamente: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generando Excel de verificación: {str(e)}")
            raise ValueError(f"Error generando archivo Excel: {str(e)}")
    
    def _configurar_estilos(self, ws):
        """Configura los estilos básicos del worksheet"""
        # Fuentes
        self.font_header = Font(name='Arial', size=12, bold=True)
        self.font_title = Font(name='Arial', size=14, bold=True)
        self.font_normal = Font(name='Arial', size=10)
        self.font_small = Font(name='Arial', size=8)
        
        # Alineación
        self.align_center = Alignment(horizontal='center', vertical='center')
        self.align_left = Alignment(horizontal='left', vertical='center')
        
        # Bordes
        self.border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Colores de fondo
        self.fill_header = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        self.fill_formula = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        self.fill_manual = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        self.fill_patron = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
    
    def _generar_encabezado(self, ws, verificacion: VerificacionMuestras):
        """Genera el encabezado del documento"""
        # Logo y título (simulado)
        ws['A1'] = "GEOFAL"
        ws['A1'].font = self.font_title
        
        ws['C1'] = "VERIFICACIÓN DE MUESTRAS CILÍNDRICAS DE CONCRETO"
        ws['C1'].font = self.font_title
        ws.merge_cells('C1:H1')
        
        # Información del documento
        ws['I1'] = f"CÓDIGO: {verificacion.codigo_documento}"
        ws['I1'].font = self.font_small
        
        ws['I2'] = f"VERSIÓN: {verificacion.version}"
        ws['I2'].font = self.font_small
        
        ws['I3'] = f"FECHA: {verificacion.fecha_documento}"
        ws['I3'].font = self.font_small
        
        ws['I4'] = f"PÁGINA: {verificacion.pagina}"
        ws['I4'].font = self.font_small
        
        # Información del verificador
        ws['A5'] = f"VERIFICADO POR: {verificacion.verificado_por or ''}"
        ws['A5'].font = self.font_normal
        
        ws['E5'] = f"FECHA VERIFIC.: {verificacion.fecha_verificacion or ''}"
        ws['E5'].font = self.font_normal
        
        # Información del cliente
        ws['A6'] = f"CLIENTE: {verificacion.cliente or ''}"
        ws['A6'].font = self.font_normal
    
    def _generar_tabla_principal(self, ws, muestras: List[MuestraVerificada]):
        """Genera la tabla principal con los datos de las muestras"""
        # Fila de inicio de la tabla
        start_row = 8
        
        # Encabezados de la tabla
        headers = [
            "N°", "Código cliente", "TIPO DE TESTIGO", "", "", "",
            "DIÁMETRO (MAYOR Y MENOR)", "", "", "",
            "PERPENDICULARIDAD (0.5°) o (1 on 100 mm)", "", "", "", "", "",
            "PLANITUD", "", "", "ACCIÓN A REALIZAR", "Conformidad Corrección realizada"
        ]
        
        # Escribir encabezados principales
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.font_header
            cell.alignment = self.align_center
            cell.fill = self.fill_header
            cell.border = self.border_thin
        
        # Encabezados de subcolumnas
        subheaders = [
            "", "", "Tipo de Testigo", "", "",
            "Diámetro 1 (mm)", "Diámetro 2 (mm)", "CUMPLE", "",
            "Máx. desviación (mm)", "", "", "", "MEDIDA <0.5°",
            "C. SUPERIOR < 0.05 mm", "C. INFERIOR < 0.05 mm", "Depresiones ≤ 5 mm",
            "(Capeo, devolución)", "N o X"
        ]
        
        for col, subheader in enumerate(subheaders, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=subheader)
            cell.font = self.font_small
            cell.alignment = self.align_center
            cell.fill = self.fill_header
            cell.border = self.border_thin
        
        # Sub-sub encabezados para CUMPLE
        ws.cell(row=start_row + 2, column=8, value="Δe >2% (mm)").font = self.font_small
        ws.cell(row=start_row + 2, column=8).alignment = self.align_center
        ws.cell(row=start_row + 2, column=8).fill = self.fill_formula
        ws.cell(row=start_row + 2, column=8).border = self.border_thin
        
        ws.cell(row=start_row + 2, column=9, value="V o X").font = self.font_small
        ws.cell(row=start_row + 2, column=9).alignment = self.align_center
        ws.cell(row=start_row + 2, column=9).fill = self.fill_formula
        ws.cell(row=start_row + 2, column=9).border = self.border_thin
        
        # Sub-sub encabezados para PERPENDICULARIDAD
        for i, p in enumerate(['P1', 'P2', 'P3', 'P4'], 10):
            ws.cell(row=start_row + 2, column=i, value=p).font = self.font_small
            ws.cell(row=start_row + 2, column=i).alignment = self.align_center
            ws.cell(row=start_row + 2, column=i).fill = self.fill_manual
            ws.cell(row=start_row + 2, column=i).border = self.border_thin
        
        ws.cell(row=start_row + 2, column=14, value="V o X").font = self.font_small
        ws.cell(row=start_row + 2, column=14).alignment = self.align_center
        ws.cell(row=start_row + 2, column=14).fill = self.fill_manual
        ws.cell(row=start_row + 2, column=14).border = self.border_thin
        
        # Sub-sub encabezados para PLANITUD
        for i, planitud in enumerate(['V o X', 'V o X', 'V o X'], 15):
            ws.cell(row=start_row + 2, column=i, value=planitud).font = self.font_small
            ws.cell(row=start_row + 2, column=i).alignment = self.align_center
            ws.cell(row=start_row + 2, column=i).fill = self.fill_patron
            ws.cell(row=start_row + 2, column=i).border = self.border_thin
        
        # Escribir datos de las muestras
        for i, muestra in enumerate(muestras, 1):
            row = start_row + 2 + i
            
            # Número de item
            ws.cell(row=row, column=1, value=i).font = self.font_normal
            ws.cell(row=row, column=1).alignment = self.align_center
            ws.cell(row=row, column=1).border = self.border_thin
            
            # Código cliente
            ws.cell(row=row, column=2, value=muestra.codigo_cliente).font = self.font_normal
            ws.cell(row=row, column=2).alignment = self.align_center
            ws.cell(row=row, column=2).border = self.border_thin
            
            # Tipo de testigo (MANUAL) - texto libre
            ws.cell(row=row, column=3, value=muestra.tipo_testigo or "").font = self.font_normal
            ws.cell(row=row, column=3).alignment = self.align_center
            ws.cell(row=row, column=3).fill = self.fill_manual
            ws.cell(row=row, column=3).border = self.border_thin
            
            # Columnas 4 y 5 vacías para mantener formato
            ws.cell(row=row, column=4, value="").font = self.font_normal
            ws.cell(row=row, column=4).alignment = self.align_center
            ws.cell(row=row, column=4).fill = self.fill_manual
            ws.cell(row=row, column=4).border = self.border_thin
            
            ws.cell(row=row, column=5, value="").font = self.font_normal
            ws.cell(row=row, column=5).alignment = self.align_center
            ws.cell(row=row, column=5).fill = self.fill_manual
            ws.cell(row=row, column=5).border = self.border_thin
            
            # Diámetros (FORMULA)
            ws.cell(row=row, column=6, value=muestra.diametro_1_mm).font = self.font_normal
            ws.cell(row=row, column=6).alignment = self.align_center
            ws.cell(row=row, column=6).fill = self.fill_formula
            ws.cell(row=row, column=6).border = self.border_thin
            
            ws.cell(row=row, column=7, value=muestra.diametro_2_mm).font = self.font_normal
            ws.cell(row=row, column=7).alignment = self.align_center
            ws.cell(row=row, column=7).fill = self.fill_formula
            ws.cell(row=row, column=7).border = self.border_thin
            
            # Tolerancia calculada
            ws.cell(row=row, column=8, value=muestra.tolerancia_porcentaje).font = self.font_normal
            ws.cell(row=row, column=8).alignment = self.align_center
            ws.cell(row=row, column=8).fill = self.fill_formula
            ws.cell(row=row, column=8).border = self.border_thin
            
            # Cumple tolerancia
            cumple_text = "✓" if muestra.cumple_tolerancia else "✗"
            ws.cell(row=row, column=9, value=cumple_text).font = self.font_normal
            ws.cell(row=row, column=9).alignment = self.align_center
            ws.cell(row=row, column=9).fill = self.fill_formula
            ws.cell(row=row, column=9).border = self.border_thin
            
            # Perpendicularidad (MANUAL)
            for j, p in enumerate([muestra.perpendicularidad_p1, muestra.perpendicularidad_p2, 
                                 muestra.perpendicularidad_p3, muestra.perpendicularidad_p4], 10):
                p_text = "✓" if p else "✗" if p is not None else ""
                ws.cell(row=row, column=j, value=p_text).font = self.font_normal
                ws.cell(row=row, column=j).alignment = self.align_center
                ws.cell(row=row, column=j).fill = self.fill_manual
                ws.cell(row=row, column=j).border = self.border_thin
            
            # Perpendicularidad cumple
            perp_cumple_text = "✓" if muestra.perpendicularidad_cumple else "✗" if muestra.perpendicularidad_cumple is not None else ""
            ws.cell(row=row, column=14, value=perp_cumple_text).font = self.font_normal
            ws.cell(row=row, column=14).alignment = self.align_center
            ws.cell(row=row, column=14).fill = self.fill_manual
            ws.cell(row=row, column=14).border = self.border_thin
            
            # Planitud (PATRON)
            planitud_values = [muestra.planitud_superior, muestra.planitud_inferior, muestra.planitud_depresiones]
            for j, planitud in enumerate(planitud_values, 15):
                planitud_text = "✓" if planitud else "✗" if planitud is not None else ""
                ws.cell(row=row, column=j, value=planitud_text).font = self.font_normal
                ws.cell(row=row, column=j).alignment = self.align_center
                ws.cell(row=row, column=j).fill = self.fill_patron
                ws.cell(row=row, column=j).border = self.border_thin
            
            # Acción a realizar (PATRON - CALCULADO)
            ws.cell(row=row, column=18, value=muestra.accion_realizar or "").font = self.font_normal
            ws.cell(row=row, column=18).alignment = self.align_center
            ws.cell(row=row, column=18).fill = self.fill_patron
            ws.cell(row=row, column=18).border = self.border_thin
            
            # Conformidad
            conf_text = "✓" if muestra.conformidad_correccion else "✗" if muestra.conformidad_correccion is not None else ""
            ws.cell(row=row, column=19, value=conf_text).font = self.font_normal
            ws.cell(row=row, column=19).alignment = self.align_center
            ws.cell(row=row, column=19).border = self.border_thin
    
    def _llenar_equipos_y_nota(self, ws, verificacion: VerificacionMuestras):
        """
        Llena la fila de equipos (fila 18) y nota (fila 19) según el template.
        
        Fila 18: B18="Código equipo", C18="Bernier", D18=código, E18="Lainas 1", F18=código, etc.
        Fila 19: A19="Nota", B19-S19=nota (línea horizontal)
        """
        # Fila 18: Equipos
        row_equipos = 18
        # B18: "Código equipo"
        ws.cell(row=row_equipos, column=2, value="Código equipo").font = self.font_normal
        ws.cell(row=row_equipos, column=2).border = self.border_thin
        
        # Equipos y códigos (C18-L18)
        equipos = [
            ("Bernier", verificacion.equipo_bernier),
            ("Lainas 1", verificacion.equipo_lainas_1),
            ("Lainas 2", verificacion.equipo_lainas_2),
            ("Escuadra", verificacion.equipo_escuadra),
            ("Balanza", verificacion.equipo_balanza)
        ]
        
        col = 3  # Empezar en C
        for nombre, codigo in equipos:
            # Nombre del equipo
            ws.cell(row=row_equipos, column=col, value=nombre).font = self.font_normal
            ws.cell(row=row_equipos, column=col).alignment = self.align_center
            ws.cell(row=row_equipos, column=col).border = self.border_thin
            col += 1
            
            # Código del equipo
            ws.cell(row=row_equipos, column=col, value=codigo or "").font = self.font_normal
            ws.cell(row=row_equipos, column=col).alignment = self.align_center
            ws.cell(row=row_equipos, column=col).border = self.border_thin
            col += 1
        
        # Fila 19: Nota
        row_nota = 19
        # A19: "Nota"
        ws.cell(row=row_nota, column=1, value="Nota").font = self.font_normal
        
        # B19-S19: Línea horizontal para la nota (si hay nota, llenar desde B hasta S)
        if verificacion.nota:
            # Llenar la nota desde B hasta S (columnas 2-19)
            nota_celdas = verificacion.nota[:200]  # Limitar longitud
            ws.cell(row=row_nota, column=2, value=nota_celdas).font = self.font_normal
            # Fusionar celdas B19-S19 para la línea
            try:
                ws.merge_cells(f'B{row_nota}:S{row_nota}')
            except:
                pass  # Si ya está fusionado, ignorar
    
    def _generar_pie_pagina(self, ws):
        """Genera el pie de página del documento"""
        # Obtener la última fila con datos
        last_row = ws.max_row
        
        # Información de contacto (fila 21 según la imagen)
        contact_info = "Web: www.geofal.com.pe / Correo: laboratorio@geofal.com.pe / Av. Marañon N°763 Los Olivos, Lima / Teléfono: 01-7543070"
        ws.cell(row=21, column=4, value=contact_info).font = self.font_small
        ws.cell(row=21, column=4).alignment = self.align_center
    
    def _ajustar_columnas(self, ws):
        """Ajusta el ancho de las columnas"""
        column_widths = {
            'A': 8,   # N°
            'B': 15,  # Código cliente
            'C': 12,  # 30x15
            'D': 12,  # 20x10
            'E': 12,  # Diamantin
            'F': 15,  # Diámetro 1
            'G': 15,  # Diámetro 2
            'H': 12,  # Δe >2%
            'I': 8,   # V o X
            'J': 8,   # P1
            'K': 8,   # P2
            'L': 8,   # P3
            'M': 8,   # P4
            'N': 12,  # MEDIDA <0.5°
            'O': 15,  # C. SUPERIOR
            'P': 15,  # C. INFERIOR
            'Q': 15,  # Depresiones
            'R': 25,  # ACCIÓN A REALIZAR
            'S': 15   # Conformidad
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _crear_excel_desde_cero(self, ws, verificacion: VerificacionMuestras):
        """
        Crea un Excel desde cero con la estructura de verificación de muestras
        """
        try:
            # Configurar estilos
            self._configurar_estilos(ws)
            
            # Título principal
            ws['A1'] = "VERIFICACIÓN DE MUESTRAS CILINDRICAS DE CONCRETO"
            ws.merge_cells('A1:Q1')
            ws['A1'].font = Font(name='Arial', size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Información del documento
            ws['Q2'] = "CÓDIGO EM-P-01.12"
            ws['Q3'] = "VERSIÓN 2"
            ws['Q4'] = f"FECHA: {datetime.now().strftime('%d/%m/%Y')}"
            ws['Q5'] = "PÁGINA: 1 de 1"
            
            # Información general
            ws['A6'] = f"CLIENTE: {verificacion.cliente or ''}"
            ws['K6'] = f"VERIFICADO POR: {verificacion.verificado_por or ''}"
            ws['Q6'] = f"FECHA VERIFIC.: {verificacion.fecha_verificacion or ''}"
            
            # Encabezados de la tabla
            ws['A8'] = "N°"
            ws['B8'] = "CÓDIGO CLIENTE"
            ws['C8'] = "TIPO DE TESTIGO"
            ws['D8'] = "DIÁMETRO (MAYOR Y MENOR)"
            ws['E8'] = "PERPENDICULARIDAD (0.5°) o (1 en 100 mm)"
            ws['F8'] = "PLANITUD"
            ws['G8'] = "ACCIÓN A REALIZAR"
            ws['H8'] = "CONFORMIDAD CORRECCIÓN"
            
            # Sub-encabezados
            ws['D9'] = "Diámetro 1"
            ws['E9'] = "Diámetro 2"
            ws['F9'] = "CUMPLE"
            ws['G9'] = "Máx. desviación (mm)"
            ws['H9'] = "ΔΦ >2% (mm)"
            ws['I9'] = "V o X"
            ws['J9'] = "P1√oX"
            ws['K9'] = "P2oX"
            ws['L9'] = "P3oX"
            ws['M9'] = "P4oX"
            ws['N9'] = "MEDIDA <0.5°"
            ws['O9'] = "SUPERIOR <0.05"
            ws['P9'] = "INFERIOR <0.05"
            ws['Q9'] = "Depresiones ≤5mm"
            ws['R9'] = "ACCIÓN"
            ws['S9'] = "V o X"
            
            # Aplicar estilos a encabezados
            for row in range(8, 10):
                for col in range(1, 20):
                    cell = ws.cell(row=row, column=col)
                    cell.font = self.font_header
                    cell.alignment = self.align_center
                    cell.fill = self.fill_header
                    cell.border = self.border_thin
            
            # Llenar datos de las muestras
            start_row = 10
            for i, muestra in enumerate(verificacion.muestras_verificadas, 1):
                row = start_row + i
                
                # Número de item
                ws.cell(row=row, column=1, value=i)
                
                # Código cliente
                ws.cell(row=row, column=2, value=muestra.codigo_cliente or "")
                
                # Tipo de testigo
                ws.cell(row=row, column=3, value=muestra.tipo_testigo or "")
                
                # Diámetros
                ws.cell(row=row, column=4, value=muestra.diametro_1_mm or "")
                ws.cell(row=row, column=5, value=muestra.diametro_2_mm or "")
                
                # Cumple tolerancia
                cumple_text = "✓" if muestra.cumple_tolerancia else "✗" if muestra.cumple_tolerancia is not None else ""
                ws.cell(row=row, column=6, value=cumple_text)
                
                # Tolerancia calculada
                ws.cell(row=row, column=7, value=muestra.tolerancia_porcentaje or "")
                
                # Tolerancia porcentaje
                ws.cell(row=row, column=8, value=muestra.tolerancia_porcentaje or "")
                
                # Cumple tolerancia (V/X)
                ws.cell(row=row, column=9, value=cumple_text)
                
                # Perpendicularidad P1, P2, P3, P4
                perp_values = [
                    muestra.perpendicularidad_p1,
                    muestra.perpendicularidad_p2,
                    muestra.perpendicularidad_p3,
                    muestra.perpendicularidad_p4
                ]
                for j, p in enumerate(perp_values, 10):
                    p_text = "✓" if p else "✗" if p is not None else ""
                    ws.cell(row=row, column=j, value=p_text)
                
                # MEDIDA <0.5°
                perp_cumple_text = "✓" if muestra.perpendicularidad_cumple else "✗" if muestra.perpendicularidad_cumple is not None else ""
                ws.cell(row=row, column=14, value=perp_cumple_text)
                
                # Planitud Superior, Inferior, Depresiones
                planitud_values = [
                    muestra.planitud_superior,
                    muestra.planitud_inferior,
                    muestra.planitud_depresiones
                ]
                for j, planitud in enumerate(planitud_values, 15):
                    planitud_text = "✓" if planitud else "✗" if planitud is not None else ""
                    ws.cell(row=row, column=j, value=planitud_text)
                
                # Acción a realizar
                ws.cell(row=row, column=18, value=muestra.accion_realizar or "")
                
                # Conformidad Corrección
                conf_text = "✓" if muestra.conformidad_correccion else "✗" if muestra.conformidad_correccion is not None else ""
                ws.cell(row=row, column=19, value=conf_text)
                
                # Aplicar estilos a la fila de datos
                for col in range(1, 20):
                    cell = ws.cell(row=row, column=col)
                    cell.font = self.font_normal
                    cell.alignment = self.align_center
                    cell.border = self.border_thin
            
            # Ajustar ancho de columnas básico
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 8
            ws.column_dimensions['J'].width = 8
            ws.column_dimensions['K'].width = 8
            ws.column_dimensions['L'].width = 8
            ws.column_dimensions['M'].width = 8
            ws.column_dimensions['N'].width = 12
            ws.column_dimensions['O'].width = 12
            ws.column_dimensions['P'].width = 12
            ws.column_dimensions['Q'].width = 12
            ws.column_dimensions['R'].width = 20
            ws.column_dimensions['S'].width = 8
            
            logger.info(f"Excel creado desde cero con {len(verificacion.muestras_verificadas)} muestras")
            
        except Exception as e:
            logger.error(f"Error creando Excel desde cero: {str(e)}")
            raise ValueError(f"Error creando Excel desde cero: {str(e)}")

    def _llenar_datos_template(self, ws, verificacion: VerificacionMuestras):
        """
        Llena los datos específicos en el template sin modificar encabezados.
        
        Args:
            ws: Worksheet de openpyxl
            verificacion: Objeto VerificacionMuestras con los datos
        """
        try:
            # Configurar estilos primero
            self._configurar_estilos(ws)
            
            # Llenar información general
            self._llenar_informacion_general(ws, verificacion)
            
            # Llenar datos de las muestras desde la fila 12
            self._llenar_datos_muestras(ws, verificacion)
            
            # Llenar equipos y nota (filas 18 y 19)
            self._llenar_equipos_y_nota(ws, verificacion)
            
            logger.info(f"Datos llenados exitosamente: {len(verificacion.muestras_verificadas)} muestras")
            
        except Exception as e:
            logger.error(f"Error llenando datos en template: {str(e)}")
            raise ValueError(f"Error llenando datos en template: {str(e)}")
    
    def _llenar_informacion_general(self, ws, verificacion: VerificacionMuestras):
        """Llena la información general del documento."""
        # Llenar campos de información general buscando las etiquetas
        campos_general = {
            "VERIFICADO POR:": verificacion.verificado_por,
            "FECHA VERIFIC.:": verificacion.fecha_verificacion,
            "CLIENTE:": verificacion.cliente
        }
        
        for etiqueta, valor in campos_general.items():
            if valor:
                self._buscar_y_llenar_celda(ws, etiqueta, valor)
    
    def _buscar_y_llenar_celda(self, ws, etiqueta: str, valor: str):
        """
        Busca una etiqueta en el worksheet y llena la celda adyacente.
        
        Args:
            ws: Worksheet de openpyxl
            etiqueta: Texto a buscar (ej: "VERIFICADO POR:")
            valor: Valor a insertar en la celda adyacente
        """
        for row in range(1, 10):
            for col in range(1, 20):
                try:
                    cell = ws.cell(row=row, column=col)
                    cell_value = cell.value
                    if cell_value and etiqueta in str(cell_value):
                        # Llenar la celda adyacente solo si está vacía
                        target_cell = ws.cell(row=row, column=col + 1)
                        if not target_cell.value:
                            target_cell.value = valor
                        break
                except Exception:
                    # Ignorar celdas fusionadas o con problemas
                    continue
    
    def _llenar_datos_muestras(self, ws, verificacion: VerificacionMuestras):
        """Llena los datos de las muestras en las filas correspondientes."""
        start_row = 12
        
        for i, muestra in enumerate(verificacion.muestras_verificadas, 1):
            row = start_row + i - 1
            self._llenar_fila_muestra(ws, row, i, muestra)
    
    def _llenar_fila_muestra(self, ws, row: int, numero: int, muestra: MuestraVerificada):
        """
        Llena una fila individual con los datos de una muestra - Formato V03.
        
        Args:
            ws: Worksheet de openpyxl
            row: Número de fila donde llenar los datos
            numero: Número de muestra
            muestra: Objeto MuestraVerificada con los datos
        """
        # Datos básicos
        self._llenar_celda_segura(ws, row, self.COLUMNS['numero'], numero)
        # Usar codigo_lem si existe, sino codigo_cliente (compatibilidad)
        codigo = getattr(muestra, 'codigo_lem', None) or getattr(muestra, 'codigo_cliente', None) or ""
        self._llenar_celda_segura(ws, row, self.COLUMNS['codigo_lem'], codigo)
        self._llenar_celda_segura(ws, row, self.COLUMNS['tipo_testigo'], muestra.tipo_testigo or "")
        
        # DIÁMETRO
        self._llenar_celda_segura(ws, row, self.COLUMNS['diametro_1'], muestra.diametro_1_mm or "")
        self._llenar_celda_segura(ws, row, self.COLUMNS['diametro_2'], muestra.diametro_2_mm or "")
        # Tolerancia porcentaje (ΔΦ 2%>)
        self._llenar_celda_segura(ws, row, self.COLUMNS['tolerancia_porcentaje'], 
                                  f"{muestra.tolerancia_porcentaje:.2f}%" if muestra.tolerancia_porcentaje else "")
        # Aceptación diámetro
        aceptacion = getattr(muestra, 'aceptacion_diametro', None) or \
                    ("Cumple" if getattr(muestra, 'cumple_tolerancia', None) else 
                     "No cumple" if getattr(muestra, 'cumple_tolerancia', None) is False else "")
        self._llenar_celda_segura(ws, row, self.COLUMNS['aceptacion_diametro'], aceptacion)
        
        # PERPENDICULARIDAD (SUP 1, SUP 2, INF 1, INF 2, MEDIDA < 0.5°)
        perp_sup1 = getattr(muestra, 'perpendicularidad_sup1', None) or getattr(muestra, 'perpendicularidad_p1', None)
        perp_sup2 = getattr(muestra, 'perpendicularidad_sup2', None) or getattr(muestra, 'perpendicularidad_p2', None)
        perp_inf1 = getattr(muestra, 'perpendicularidad_inf1', None) or getattr(muestra, 'perpendicularidad_p3', None)
        perp_inf2 = getattr(muestra, 'perpendicularidad_inf2', None) or getattr(muestra, 'perpendicularidad_p4', None)
        perp_medida = getattr(muestra, 'perpendicularidad_medida', None) or getattr(muestra, 'perpendicularidad_cumple', None)
        
        self._llenar_celda_segura(ws, row, self.COLUMNS['perpendicularidad_sup1'], 
                                  self._formatear_aceptacion(perp_sup1))
        self._llenar_celda_segura(ws, row, self.COLUMNS['perpendicularidad_sup2'], 
                                  self._formatear_aceptacion(perp_sup2))
        self._llenar_celda_segura(ws, row, self.COLUMNS['perpendicularidad_inf1'], 
                                  self._formatear_aceptacion(perp_inf1))
        self._llenar_celda_segura(ws, row, self.COLUMNS['perpendicularidad_inf2'], 
                                  self._formatear_aceptacion(perp_inf2))
        # MEDIDA < 0.5° está en PERPENDICULARIDAD (columna L)
        self._llenar_celda_segura(ws, row, self.COLUMNS['perpendicularidad_medida'], 
                                  self._formatear_aceptacion(perp_medida))
        
        # PLANITUD (sin fusionar: M, N, O)
        planitud_sup = getattr(muestra, 'planitud_superior_aceptacion', None) or \
                      (self._formatear_aceptacion(getattr(muestra, 'planitud_superior', None)))
        planitud_inf = getattr(muestra, 'planitud_inferior_aceptacion', None) or \
                      (self._formatear_aceptacion(getattr(muestra, 'planitud_inferior', None)))
        planitud_dep = getattr(muestra, 'planitud_depresiones_aceptacion', None) or \
                      (self._formatear_aceptacion(getattr(muestra, 'planitud_depresiones', None)))
        
        # C.SUPERIOR (M)
        self._llenar_celda_segura(ws, row, self.COLUMNS['planitud_superior'], planitud_sup)
        # C.INFERIOR (N)
        self._llenar_celda_segura(ws, row, self.COLUMNS['planitud_inferior'], planitud_inf)
        # DEPRESIONES (O)
        self._llenar_celda_segura(ws, row, self.COLUMNS['planitud_depresiones'], planitud_dep)
        
        # ACCIÓN A REALIZAR (P)
        self._llenar_celda_segura(ws, row, self.COLUMNS['accion'], muestra.accion_realizar or "")
        
        # CONFORMIDAD (Q)
        conformidad = getattr(muestra, 'conformidad', None) or \
                     (self._formatear_checkbox(getattr(muestra, 'conformidad_correccion', None)))
        self._llenar_celda_segura(ws, row, self.COLUMNS['conformidad'], conformidad)
        
        # LONGITUD (L/D ≤1.75) - R, S, T
        self._llenar_celda_segura(ws, row, self.COLUMNS['longitud_1'], muestra.longitud_1_mm or "")
        self._llenar_celda_segura(ws, row, self.COLUMNS['longitud_2'], muestra.longitud_2_mm or "")
        self._llenar_celda_segura(ws, row, self.COLUMNS['longitud_3'], muestra.longitud_3_mm or "")
        
        # MASA MUESTRA AIRE (U) - dígitos
        self._llenar_celda_segura(ws, row, self.COLUMNS['masa'], muestra.masa_muestra_aire_g or "")
        
        # PESAR / NO PESAR (V)
        pesar = getattr(muestra, 'pesar', None) or ""
        self._llenar_celda_segura(ws, row, self.COLUMNS['pesar'], pesar)
    
    def _llenar_celda_segura(self, ws, row: int, col: int, valor):
        """
        Llena una celda de forma segura, evitando errores con celdas fusionadas.
        
        Args:
            ws: Worksheet de openpyxl
            row: Número de fila
            col: Número de columna
            valor: Valor a insertar
        """
        try:
            ws.cell(row=row, column=col, value=valor)
        except Exception:
            # Ignorar errores con celdas fusionadas
            pass
    
    def _formatear_checkbox(self, valor: bool) -> str:
        """
        Convierte un valor booleano en el formato de checkbox del Excel.
        
        Args:
            valor: Valor booleano (puede ser None)
            
        Returns:
            str: "✓" si True, "✗" si False, "" si None
        """
        if valor is True:
            return "✓"
        elif valor is False:
            return "✗"
        else:
            return ""
    
    def _formatear_aceptacion(self, valor) -> str:
        """
        Convierte un valor booleano en formato de aceptación (Cumple/No cumple).
        
        Args:
            valor: Valor booleano o string (puede ser None)
            
        Returns:
            str: "Cumple" si True, "No cumple" si False, valor original si es string, "" si None
        """
        if isinstance(valor, str):
            return valor
        elif valor is True:
            return "Cumple"
        elif valor is False:
            return "No cumple"
        else:
            return ""
