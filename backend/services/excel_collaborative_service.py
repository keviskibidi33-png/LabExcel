import io
from copy import copy
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font

class ExcelCollaborativeService:
    """Servicio para modificar archivos Excel existentes con datos del formulario"""
    
    # Constantes para el template
    TEMPLATE_PATH = "templates/recepcion_template.xlsx"
    FILA_INICIO_MUESTRAS = 23
    FILA_FOOTER_ORIGINAL = 42
    MAX_ITEMS_SIN_EXPANSION = 17
    ALTURA_FILA_57 = 30.0
    ALTURA_FILA_ESTANDAR = 15.0
    
    def __init__(self):
        self.template_path = self.TEMPLATE_PATH
    
    def modificar_excel_con_datos(self, recepcion_data: Dict[str, Any], muestras: List[Dict[str, Any]], 
                                 template_path: Optional[str] = None) -> bytes:
        """Modificar archivo Excel existente con datos del formulario"""
        
        # Ruta del template
        template_file = template_path or self.template_path
        # print(f"USANDO TEMPLATE: {template_file}")
        
        # Cargar template directamente
        workbook = openpyxl.load_workbook(template_file)
        worksheet = workbook.active
        
        # Guardar número de items para uso en safe_set_cell
        self._total_items = len(muestras)
        
        # Cambiar "X" por "Descripción" - manejar celdas fusionadas
        try:
            # Buscar la celda que contiene "X" en la fila 22
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                cell_ref = f"{col}22"
                cell_value = worksheet[cell_ref].value
                if cell_value and 'X' in str(cell_value):
                    # Verificar si está en un rango fusionado
                    for merged_range in worksheet.merged_cells.ranges:
                        if cell_ref in merged_range:
                            # Usar la celda superior izquierda del rango fusionado
                            top_left = merged_range.min_row, merged_range.min_col
                            target_cell = worksheet.cell(row=top_left[0], column=top_left[1])
                            if target_cell.value:
                                # Cambiar "X" por "Descripción" en el texto fusionado
                                new_value = str(target_cell.value).replace('X', 'Descripción')
                                target_cell.value = new_value
                                # Cambio aplicado en rango fusionado
                            break
                    else:
                        # Si no está fusionada, cambiar directamente
                        worksheet[cell_ref].value = str(cell_value).replace('X', 'Descripción')
                        # Cambio aplicado en celda no fusionada
                    break
        except Exception:
            # Evitar interrumpir flujo por ajustes de encabezado
            pass
        
        # Rellenar datos
        self._rellenar_datos_recepcion(worksheet, recepcion_data)
        self._rellenar_datos_muestras(worksheet, muestras)
        self._ajustar_ancho_columnas(worksheet)
        
        # Guardar
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # print("TEMPLATE REAL USADO EXITOSAMENTE")
        return excel_buffer.getvalue()
    
    def _rellenar_datos_recepcion(self, worksheet, recepcion_data: Dict[str, Any]):
        """Rellenar datos de la recepción en el Excel"""
        
        def safe_set_cell(cell_ref, value):
            try:
                if isinstance(value, str):
                    value = value.strip()
                
                cell = worksheet[cell_ref]
                target_cell = cell
                
                # Verificar si es una celda fusionada
                for merged_range in worksheet.merged_cells.ranges:
                    if cell_ref in merged_range:
                        top_left = merged_range.min_row, merged_range.min_col
                        target_cell = worksheet.cell(row=top_left[0], column=top_left[1])
                        break
                
                target_cell.value = value
                
                # Centrar las X en los cuadros de checkboxes
                if cell_ref in ['B46', 'B47'] and value == 'X':
                    target_cell.alignment = Alignment(horizontal='center', vertical='center')
                # Aplicar wrap_text solo para campos específicos y solo cuando hay muchos items
                elif cell_ref in ['H46', 'D49', 'H49']:  # fecha estimada, entregado por, recibido por
                    # Solo aplicar wrap_text si hay muchos items (más de 17)
                    if hasattr(self, '_total_items') and self._total_items > 17:
                        target_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        # Asegurar que la fila tenga altura suficiente para el wrap text
                        worksheet.row_dimensions[target_cell.row].height = max(worksheet.row_dimensions[target_cell.row].height, 30)
                    else:
                        # Mantener alineación original del template
                        target_cell.alignment = Alignment(horizontal='center', vertical='bottom')
                else:
                    # Solo aplicar wrap_text si hay muchos items
                    if hasattr(self, '_total_items') and self._total_items > 17:
                        target_cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
                    else:
                        target_cell.alignment = Alignment(horizontal='left', vertical='bottom')
                # Mantener altura original del template
                
            except Exception:
                # No interrumpir por errores de formato puntuales
                pass
        
        # Datos principales
        safe_set_cell('D6', recepcion_data.get('numero_recepcion', ''))
        safe_set_cell('D7', recepcion_data.get('numero_cotizacion', ''))
        # D9 eliminado - no poner "SOLICITO EJECUCIÓN DE ENSAYOS"
        safe_set_cell('J6', recepcion_data.get('fecha_recepcion', ''))
        safe_set_cell('J7', recepcion_data.get('numero_ot', ''))
        # I8 eliminado - no poner código de trazabilidad
        
        # Datos del cliente
        safe_set_cell('D10', recepcion_data.get('cliente', ''))
        safe_set_cell('D11', recepcion_data.get('domicilio_legal', ''))
        safe_set_cell('D12', recepcion_data.get('ruc', ''))
        safe_set_cell('D13', recepcion_data.get('persona_contacto', ''))
        safe_set_cell('D14', recepcion_data.get('email', ''))
        safe_set_cell('H14', recepcion_data.get('telefono', ''))
        
        # Datos para el informe
        safe_set_cell('D16', recepcion_data.get('solicitante', ''))
        safe_set_cell('D17', recepcion_data.get('domicilio_solicitante', ''))
        safe_set_cell('D18', recepcion_data.get('proyecto', ''))
        safe_set_cell('D19', recepcion_data.get('ubicacion', ''))
        
        # Fecha estimada
        safe_set_cell('H46', recepcion_data.get('fecha_estimada_culminacion', ''))
        
        # Emisión - Las X van en columna B donde están los cuadros
        # Limpiar primero las celdas (eliminar cualquier X predefinida del template)
        worksheet['B46'].value = None
        worksheet['B47'].value = None
        
        # Aplicar lógica real (no mock predefinido)
        if recepcion_data.get('emision_fisica', False):
            safe_set_cell('B46', 'X')
        if recepcion_data.get('emision_digital', False):
            safe_set_cell('B47', 'X')
        
        # No agregar X en D22 - esa columna es para códigos de muestras
        
        # Entregado/Recibido - Nuevas celdas sin fusionar
        safe_set_cell('D49', recepcion_data.get('entregado_por', ''))
        safe_set_cell('H49', recepcion_data.get('recibido_por', ''))
        
        # Información de contacto en columna D
        self._agregar_informacion_contacto(worksheet)
        
        print("Datos de recepción rellenados")
    
    def _rellenar_datos_muestras(self, worksheet, muestras: List[Dict[str, Any]]):
        """Rellenar datos de las muestras manteniendo el footer en su posición."""

        def safe_set_cell(cell_ref: str, value: Any) -> None:
            try:
                if isinstance(value, str):
                    value = value.strip()

                cell = worksheet[cell_ref]
                destino = cell

                for merged_range in worksheet.merged_cells.ranges:
                    if cell_ref in merged_range:
                        destino = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        break

                borde = destino.border
                relleno = destino.fill
                fuente = destino.font
                alineacion = destino.alignment

                destino.value = value
                destino.border = borde
                destino.fill = relleno
                destino.font = fuente
                destino.alignment = alineacion
            except Exception:
                pass

        columnas_tabla = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

        fila_inicio = self.FILA_INICIO_MUESTRAS
        total_items = len(muestras)
        altura_item = worksheet.row_dimensions[fila_inicio].height or self.ALTURA_FILA_ESTANDAR

        footer_row = self._find_footer_row(worksheet)
        if not footer_row:
            raise ValueError("No se encontró el footer en la plantilla original")

        footer_row = self._asegurar_capacidad_items(worksheet, footer_row, total_items)
        self._unmerge_item_area(worksheet, footer_row)
        
        # Ajustar ancho de columna A para evitar "#" en números
        self._ajustar_ancho_columna_a(worksheet, total_items)
        
        # SIEMPRE asegurar que los campos importantes estén presentes
        self._asegurar_campos_importantes(worksheet)
        
        # Lógica dinámica basada en el número de items
        if total_items >= 40:
            # Para 40+ items: aplicar todas las optimizaciones
            self._refusionar_items_con_control(worksheet, fila_inicio, total_items)
            self._fusionar_celdas_footer(worksheet)
            print(f"Aplicando optimizaciones completas para {total_items} items")
        elif total_items > 17:
            # Para 18-39 items: solo fusionar footer, mantener template original
            self._fusionar_celdas_footer(worksheet)
            print(f"Aplicando solo fusión de footer para {total_items} items")
        else:
            # Para 17 o menos items: mantener template original
            print(f"Manteniendo template original para {total_items} items")

        for indice, muestra in enumerate(muestras):
            fila_actual = fila_inicio + indice
            worksheet.row_dimensions[fila_actual].height = altura_item

            # Forzar número de item correcto
            item_numero = indice + 1
            print(f"Fila {fila_actual}: item_numero forzado={item_numero}")
            
            # Debug específico para celda A
            try:
                celda_a = worksheet[f'A{fila_actual}']
                print(f"  Celda A{fila_actual} antes: {celda_a.value}")
                celda_a.value = item_numero
                print(f"  Celda A{fila_actual} después: {celda_a.value}")
            except Exception as e:
                print(f"  Error en celda A{fila_actual}: {e}")
            
            safe_set_cell(f'A{fila_actual}', item_numero)
            codigo_ref = f'B{fila_actual}'
            safe_set_cell(codigo_ref, muestra.get('codigo_muestra_lem', ''))
            try:
                worksheet[codigo_ref].number_format = '@'
            except Exception:
                pass
            safe_set_cell(f'D{fila_actual}', muestra.get('identificacion_muestra', ''))
            safe_set_cell(f'E{fila_actual}', muestra.get('estructura', ''))
            safe_set_cell(f'F{fila_actual}', muestra.get('fc_kg_cm2', ''))
            safe_set_cell(f'G{fila_actual}', muestra.get('fecha_moldeo', ''))
            safe_set_cell(f'H{fila_actual}', muestra.get('hora_moldeo', ''))
            safe_set_cell(f'I{fila_actual}', muestra.get('edad', ''))
            safe_set_cell(f'J{fila_actual}', muestra.get('fecha_rotura', ''))
            safe_set_cell(f'K{fila_actual}', 'SI' if muestra.get('requiere_densidad', False) else 'NO')
            self._merge_item_row(worksheet, fila_actual)

        self._limpiar_filas_restantes(worksheet, fila_inicio + total_items, footer_row, columnas_tabla)
        # Solo eliminar segunda línea web si hay muchos items
        if total_items > 17:
            self._eliminar_segunda_linea_web(worksheet)
        
        # Centrar datos de filas específicas (20, 21, 29) cuando tengan items
        self._centrar_filas_especificas(worksheet, fila_inicio, total_items)
        
        # SIEMPRE centrar el último número de item
        self._centrar_ultimo_item(worksheet, fila_inicio, total_items)
    
    def _centrar_filas_especificas(self, worksheet, fila_inicio: int, total_items: int) -> None:
        """Centrar datos de filas específicas (items 20, 21, 24, 25, 26, 27, 29, 40) cuando tengan items"""
        from openpyxl.styles import Alignment
        
        # Items específicos a centrar (20, 21, 24, 25, 26, 27, 29, 40)
        items_a_centrar = [20, 21, 24, 25, 26, 27, 29, 40]
        
        for item_num in items_a_centrar:
            # Convertir número de item a fila del Excel
            # Los items empiezan en fila_inicio, entonces item 20 = fila_inicio + 19
            fila_excel = fila_inicio + (item_num - 1)
            
            # Verificar si la fila está dentro del rango de items
            if fila_inicio <= fila_excel < fila_inicio + total_items:
                print(f"Centrando fila {fila_excel} (item {item_num})")
                
                # Aplicar alineación centrada a todas las columnas de datos
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                    try:
                        celda = worksheet[f'{col}{fila_excel}']
                        
                        # Verificar si la celda está en un rango fusionado
                        celda_destino = celda
                        for merged_range in worksheet.merged_cells.ranges:
                            if f'{col}{fila_excel}' in merged_range:
                                celda_destino = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                                break
                        
                        # Aplicar alineación centrada
                        celda_destino.alignment = Alignment(horizontal='center', vertical='center')
                        print(f"  Centrada celda {col}{fila_excel}")
                        
                    except Exception as e:
                        print(f"  Error centrando {col}{fila_excel}: {e}")
                        pass
    
    def _ajustar_ancho_columnas(self, worksheet):
        """Ajustar el ancho de las columnas"""
        try:
            # Volver a los anchos originales - no modificar B, C y G (mantener template original)
            worksheet.column_dimensions['D'].width = 30
            worksheet.column_dimensions['H'].width = 15
            worksheet.column_dimensions['I'].width = 20
            worksheet.column_dimensions['J'].width = 12
            # G mantiene su ancho original del template
            
            # No ajustar altura de fila 8 - código de trazabilidad eliminado
            
            # print("Columnas ajustadas a anchos originales")
            
        except Exception:
            # No detener por ajustes visuales
            pass

    def _asegurar_capacidad_items(self, worksheet, footer_row: int, total_items: int) -> int:
        capacidad_actual = footer_row - self.FILA_INICIO_MUESTRAS
        if total_items <= capacidad_actual:
            return footer_row

        filas_extra = total_items - capacidad_actual
        fila_patron = footer_row - 2
        fila_separador = footer_row - 1

        worksheet.insert_rows(fila_separador, amount=filas_extra)

        for offset in range(filas_extra):
            fila_destino = fila_separador + offset
            self._copiar_estilo_fila(worksheet, fila_patron, fila_destino)
            self._clonar_fusiones_fila(worksheet, fila_patron, fila_destino)

        return footer_row + filas_extra

    def _copiar_estilo_fila(self, worksheet, fila_origen: int, fila_destino: int) -> None:
        worksheet.row_dimensions[fila_destino].height = worksheet.row_dimensions[fila_origen].height

        columnas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        for columna in columnas:
            celda_origen = worksheet[f'{columna}{fila_origen}']
            celda_destino = worksheet[f'{columna}{fila_destino}']
            celda_destino.value = None
            celda_destino.font = copy(celda_origen.font)
            celda_destino.border = copy(celda_origen.border)
            celda_destino.fill = copy(celda_origen.fill)
            celda_destino.number_format = celda_origen.number_format
            celda_destino.protection = copy(celda_origen.protection)
            celda_destino.alignment = copy(celda_origen.alignment)
        self._merge_item_row(worksheet, fila_destino)

    def _clonar_fusiones_fila(self, worksheet, fila_origen: int, fila_destino: int) -> None:
        rangos_existentes = {r.coord for r in worksheet.merged_cells.ranges}

        for rango in list(worksheet.merged_cells.ranges):
            if rango.min_row == rango.max_row == fila_origen:
                celda_inicio = worksheet.cell(row=fila_destino, column=rango.min_col)
                celda_fin = worksheet.cell(row=fila_destino, column=rango.max_col)
                coord = f"{celda_inicio.coordinate}:{celda_fin.coordinate}"
                if coord not in rangos_existentes:
                    worksheet.merge_cells(coord)
                    rangos_existentes.add(coord)
        
        # Asegurar bordes después de clonar fusiones
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar bordes a todas las celdas de la fila destino
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            try:
                worksheet[f'{col}{fila_destino}'].border = thin_border
            except Exception:
                pass

    def _limpiar_filas_restantes(self, worksheet, fila_inicio_libre: int, fila_footer: int, columnas: List[str]) -> None:
        """Limpiar filas restantes manteniendo el formato correcto"""
        for fila in range(fila_inicio_libre, fila_footer):
            # No limpiar filas críticas del footer (últimas 10 filas)
            if fila < fila_footer - 10:  # Proteger más filas del footer
                for columna in columnas:
                    referencia = f'{columna}{fila}'
                    try:
                        celda = worksheet[referencia]
                        destino = celda
                        for rango in worksheet.merged_cells.ranges:
                            if referencia in rango:
                                destino = worksheet.cell(row=rango.min_row, column=rango.min_col)
                                break
                        destino.value = ""
                    except Exception:
                        continue
                # Solo aplicar merge si no es una fila crítica
                if fila < fila_footer - 15:  # Proteger aún más filas
                    self._merge_item_row(worksheet, fila)

    def _find_footer_row(self, worksheet) -> Optional[int]:
        for fila in range(1, worksheet.max_row + 1):
            valor = worksheet.cell(row=fila, column=1).value
            if isinstance(valor, str) and '(1) OBLIGATORIO' in valor:
                return fila
        return None

    def _unmerge_item_area(self, worksheet, footer_row: int) -> None:
        inicio = self.FILA_INICIO_MUESTRAS
        fin = footer_row - 1
        to_unmerge = []
        for rango in list(worksheet.merged_cells.ranges):
            if rango.max_row >= inicio and rango.min_row <= fin:
                to_unmerge.append(rango.coord)
        for coord in to_unmerge:
            worksheet.unmerge_cells(coord)

    def _refusionar_items_con_control(self, worksheet, fila_inicio: int, total_items: int) -> None:
        """Re-fusionar items con control total sobre bordes y alineación"""
        from openpyxl.styles import Border, Side, Alignment
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for indice in range(total_items):
            fila_actual = fila_inicio + indice
            
            # Re-fusionar B:C con control total
            try:
                worksheet.merge_cells(f'B{fila_actual}:C{fila_actual}')
            except Exception:
                pass
            
            # Aplicar bordes y alineación a todas las celdas
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                try:
                    celda = worksheet[f'{col}{fila_actual}']
                    celda.border = thin_border
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                except Exception:
                    pass

    @staticmethod
    def _merge_item_row(worksheet, fila: int) -> None:
        coord = f"B{fila}:C{fila}"
        for rango in worksheet.merged_cells.ranges:
            if rango.coord == coord:
                return
        worksheet.merge_cells(coord)
        
        # Asegurar que los bordes se mantengan después de la fusión
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar bordes a todas las celdas de la fila
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            try:
                worksheet[f'{col}{fila}'].border = thin_border
            except Exception:
                pass

    def _eliminar_segunda_linea_web(self, worksheet) -> None:
        """Elimina el texto duplicado con la información de Web en el footer, pero mantiene la información de contacto."""

        ocurrencias = 0
        for row in range(1, worksheet.max_row + 1):
            valor = worksheet.cell(row=row, column=1).value
            if isinstance(valor, str) and "Web: www.geofal.com.pe" in valor:
                ocurrencias += 1
                # Solo eliminar si es realmente duplicado (más de 2 ocurrencias)
                if ocurrencias > 2:
                    celda = worksheet.cell(row=row, column=1)
                    coordenada = celda.coordinate
                    celda_destino = celda
                    for merged_range in worksheet.merged_cells.ranges:
                        if coordenada in merged_range:
                            celda_destino = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                            break
                    # Solo eliminar duplicados reales, mantener información de contacto principal
                    if "Web: www.geofal.com.pe" in str(celda_destino.value):
                        celda_destino.value = ""
                    break

    def _agregar_informacion_contacto(self, worksheet) -> None:
        """Agregar información de contacto en columna D"""
        # Buscar la fila donde está la información de contacto
        for row in range(50, worksheet.max_row + 1):
            valor = worksheet.cell(row=row, column=1).value
            if isinstance(valor, str) and "Web:" in valor:
                # Encontrar la celda D de esa fila
                celda_d = worksheet.cell(row=row, column=4)  # Columna D
                celda_d.value = "Web: www.geofal.com.pe / Correo: laboratorio@geofal.com.pe / Av. Marañon N°763 Los Olivos, Lima / Teléfono: 01-7543070"
                break

    def _fusionar_celdas_footer(self, worksheet) -> None:
        """Fusionar celdas del footer para evitar texto cortado con muchos items"""
        # Buscar la fila del footer buscando "Entregado por:" o "Recibido por:"
        footer_row = None
        for row in range(65, worksheet.max_row + 1):
            valor_a = worksheet.cell(row=row, column=1).value
            valor_b = worksheet.cell(row=row, column=2).value
            
            # Buscar por "Entregado por:" o "Recibido por:"
            if isinstance(valor_a, str) and ("Entregado por:" in valor_a or "Recibido por:" in valor_a):
                footer_row = row
                print(f"Encontrada fila del footer por texto en columna A: {row}")
                break
            elif isinstance(valor_b, str) and ("Entregado por:" in valor_b or "Recibido por:" in valor_b):
                footer_row = row
                print(f"Encontrada fila del footer por texto en columna B: {row}")
                break
            # Fallback: buscar por "Web:" o "geofal"
            elif isinstance(valor_a, str) and ("Web:" in valor_a or "geofal" in valor_a.lower()):
                footer_row = row
                print(f"Encontrada fila del footer por Web en columna A: {row}")
                break
        
        if footer_row:
            print(f"Encontrada fila del footer: {footer_row}")
            # Fusionar A:B y F:G en la fila del footer
            try:
                # Verificar si ya están fusionadas
                merged_ranges = [r.coord for r in worksheet.merged_cells.ranges]
                
                if f'A{footer_row}:B{footer_row}' not in merged_ranges:
                    worksheet.merge_cells(f'A{footer_row}:B{footer_row}')
                    print(f"Fusionada A:B en fila {footer_row}")
                
                if f'F{footer_row}:G{footer_row}' not in merged_ranges:
                    worksheet.merge_cells(f'F{footer_row}:G{footer_row}')
                    print(f"Fusionada F:G en fila {footer_row}")
                
                # Solo aplicar altura y wrap_text si hay muchos items (40+)
                if hasattr(self, '_total_items') and self._total_items >= 40:
                    worksheet.row_dimensions[footer_row].height = 30  # Altura para dos líneas de texto
                    
                    # Aplicar wrap_text y alineación a las celdas relevantes
                    from openpyxl.styles import Alignment
                    worksheet.cell(row=footer_row, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    worksheet.cell(row=footer_row, column=6).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    print(f"Aplicando altura y wrap_text para {self._total_items} items")
                else:
                    print(f"Manteniendo altura original del template para {getattr(self, '_total_items', 'desconocido')} items")
                    
            except Exception as e:
                print(f"Error fusionando footer en fila {footer_row}: {e}")
        else:
            print("No se encontró la fila del footer")
            # Intentar fusionar directamente en la fila 70
            try:
                print("Intentando fusionar directamente en fila 70...")
                worksheet.merge_cells('A70:B70')
                worksheet.merge_cells('F70:G70')
                print("Fusionadas celdas en fila 70")
                
                # También establecer altura y wrap_text para la fila 70 directa
                worksheet.row_dimensions[70].height = 30
                from openpyxl.styles import Alignment
                worksheet.cell(row=70, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                worksheet.cell(row=70, column=6).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            except Exception as e:
                print(f"Error fusionando fila 70: {e}")

    def _ajustar_ancho_columna_a(self, worksheet, total_items: int) -> None:
        """Ajustar ancho de columna A de manera precisa para evitar '#' en números"""
        # Calcular ancho necesario basado en el número de items
        if total_items <= 9:
            # Para 1-9: ancho mínimo
            ancho = 6.0
        elif total_items <= 99:
            # Para 10-99: ancho medio
            ancho = 8.0
        else:
            # Para 100+: ancho máximo
            ancho = 10.0
        
        # Aplicar el ancho con precisión de 2-3px
        worksheet.column_dimensions['A'].width = ancho
        print(f"Ajustado ancho de columna A a {ancho} para {total_items} items")

    def _asegurar_campos_importantes(self, worksheet) -> None:
        """Asegurar que los campos importantes SIEMPRE estén presentes sin importar el número de items"""
        # Buscar y asegurar que "Entregado por:" y "Recibido por:" estén presentes
        for row in range(45, worksheet.max_row + 1):
            valor_a = worksheet.cell(row=row, column=1).value
            valor_b = worksheet.cell(row=row, column=2).value
            
            # Si encontramos "Entregado por:" o "Recibido por:", asegurar que estén completos
            if isinstance(valor_a, str) and ("Entregado por:" in valor_a or "Recibido por:" in valor_a):
                print(f"Campo importante encontrado en fila {row}: {valor_a}")
                # Asegurar que la celda tenga el texto completo
                if "Entregado por:" in valor_a and not valor_a.strip().endswith(":"):
                    worksheet.cell(row=row, column=1).value = "Entregado por:"
                elif "Recibido por:" in valor_a and not valor_a.strip().endswith(":"):
                    worksheet.cell(row=row, column=1).value = "Recibido por:"
                break
            elif isinstance(valor_b, str) and ("Entregado por:" in valor_b or "Recibido por:" in valor_b):
                print(f"Campo importante encontrado en columna B fila {row}: {valor_b}")
                # Asegurar que la celda tenga el texto completo
                if "Entregado por:" in valor_b and not valor_b.strip().endswith(":"):
                    worksheet.cell(row=row, column=2).value = "Entregado por:"
                elif "Recibido por:" in valor_b and not valor_b.strip().endswith(":"):
                    worksheet.cell(row=row, column=2).value = "Recibido por:"
                break
        
        print("Campos importantes verificados y asegurados")

    def _centrar_ultimo_item(self, worksheet, fila_inicio: int, total_items: int) -> None:
        """SIEMPRE centrar el último número de item sin importar la cantidad"""
        from openpyxl.styles import Alignment
        
        if total_items > 0:
            # Calcular la fila del último item
            ultima_fila = fila_inicio + total_items - 1
            
            # Centrar el número en la columna A
            try:
                celda_a = worksheet[f'A{ultima_fila}']
                celda_a.alignment = Alignment(horizontal='center', vertical='center')
                print(f"Centrado último item en fila {ultima_fila} (item {total_items})")
            except Exception as e:
                print(f"Error centrando último item: {e}")




