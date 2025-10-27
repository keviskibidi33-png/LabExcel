"""
Servicio para generación de archivos Excel de Control de Concreto
Basado en el patrón de recepción de muestras con funcionalidad de búsqueda inteligente
"""

import os
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ConcretoExcelService:
    """Servicio para generar archivos Excel de Control de Concreto con búsqueda inteligente"""
    
    # Configuración del template
    TEMPLATE_PATH = "templates/concreto/CONTROL CONCRETO.xlsx"
    FILA_INICIO_DATOS = 8  # Fila donde empiezan los datos (después del header)
    ALTURA_FILA_ESTANDAR = 20.0
    
    def __init__(self):
        self.template_path = os.path.join(os.path.dirname(__file__), '..', self.TEMPLATE_PATH)
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template no encontrado: {self.template_path}")
    
    def generar_excel_concreto(self, datos_concreto: List[Dict[str, Any]], 
                             datos_cliente: Optional[Dict[str, Any]] = None) -> str:
        """
        Generar archivo Excel de Control de Concreto con búsqueda inteligente
        
        Args:
            datos_concreto: Lista de datos de probetas de concreto
            datos_cliente: Datos del cliente para relleno automático (opcional)
        
        Returns:
            Ruta del archivo generado
        """
        try:
            # Cargar template
            workbook = load_workbook(self.template_path)
            worksheet = workbook.active
            
            logger.info(f"Generando Excel de Control de Concreto con {len(datos_concreto)} probetas")
            
            # Rellenar datos del cliente si están disponibles
            if datos_cliente:
                self._rellenar_datos_cliente(worksheet, datos_cliente)
            
            # Rellenar datos de probetas
            self._rellenar_datos_probetas(worksheet, datos_concreto)
            
            # Aplicar formato final
            self._aplicar_formato_final(worksheet, len(datos_concreto))
            
            # Guardar archivo
            output_path = self._guardar_archivo(workbook, "control_concreto")
            logger.info(f"Archivo generado exitosamente: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error generando Excel de concreto: {e}")
            raise
    
    def _rellenar_datos_cliente(self, worksheet, datos_cliente: Dict[str, Any]) -> None:
        """Rellenar datos del cliente en el header del documento"""
        try:
            # Rellenar información del documento en la fila 2 (H2:J2) - NO en la fila 7
            if 'codigo_documento' in datos_cliente:
                worksheet['H2'].value = datos_cliente['codigo_documento']
            
            if 'version' in datos_cliente:
                worksheet['H3'].value = datos_cliente['version']
            
            if 'fecha_documento' in datos_cliente:
                worksheet['H4'].value = datos_cliente['fecha_documento']
            
            if 'pagina' in datos_cliente:
                worksheet['H5'].value = datos_cliente['pagina']
            
            logger.info("Datos del cliente rellenados en el header (fila 2-5)")
            
        except Exception as e:
            logger.error(f"Error rellenando datos del cliente: {e}")
    
    def _rellenar_datos_probetas(self, worksheet, probetas: List[Dict[str, Any]]) -> None:
        """Rellenar datos de las probetas de concreto"""
        
        def safe_set_cell(cell_ref: str, value: Any) -> None:
            try:
                if isinstance(value, str):
                    value = value.strip()
                
                cell = worksheet[cell_ref]
                destino = cell
                
                # Verificar si la celda está fusionada
                for merged_range in worksheet.merged_cells.ranges:
                    if cell_ref in merged_range:
                        destino = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        break
                
                destino.value = value
                
            except Exception as e:
                logger.error(f"Error estableciendo celda {cell_ref}: {e}")
        
        fila_inicio = self.FILA_INICIO_DATOS
        altura_item = self.ALTURA_FILA_ESTANDAR
        
        for indice, probeta in enumerate(probetas):
            fila_actual = fila_inicio + indice
            worksheet.row_dimensions[fila_actual].height = altura_item
            
            # ITEM (columna A)
            item_numero = indice + 1
            safe_set_cell(f'A{fila_actual}', item_numero)
            
            # ORDEN TRABAJO (columna B)
            orden_trabajo = probeta.get('orden_trabajo', '')
            if not orden_trabajo and 'codigo_muestra' in probeta:
                # Búsqueda inteligente: extraer orden de trabajo del código de muestra
                codigo = probeta['codigo_muestra']
                if '-' in codigo:
                    partes = codigo.split('-')
                    if len(partes) >= 3:
                        orden_trabajo = f"{partes[0]}-{partes[1]}-{partes[2]}"
            safe_set_cell(f'B{fila_actual}', orden_trabajo)
            
            # CÓDIGO MUESTRA LEM (columna C)
            safe_set_cell(f'C{fila_actual}', probeta.get('codigo_muestra', ''))
            
            # Clientes (columna D)
            safe_set_cell(f'D{fila_actual}', probeta.get('codigo_muestra_cliente', ''))
            
            # FECHA ROTURA PROGRAMADO (columna E)
            safe_set_cell(f'E{fila_actual}', probeta.get('fecha_rotura', ''))
            
            # Elemento (columna F) - Dropdown con opciones: 4in x 8in, 6in x 12in, cubos, viga
            elemento = probeta.get('elemento', '')
            safe_set_cell(f'F{fila_actual}', elemento)
            
            # F'C (columna G)
            fc_value = probeta.get('fc_kg_cm2', '')
            if fc_value:
                try:
                    fc_numeric = float(fc_value)
                    safe_set_cell(f'G{fila_actual}', fc_numeric)
                except (ValueError, TypeError):
                    safe_set_cell(f'G{fila_actual}', fc_value)
            else:
                safe_set_cell(f'G{fila_actual}', '')
            
            # STATUS ENSAYADO (columna H)
            status = probeta.get('status_ensayado', 'PENDIENTE')
            safe_set_cell(f'H{fila_actual}', status)
            
            # Aplicar formato a la fila
            self._aplicar_formato_fila(worksheet, fila_actual)
        
        logger.info(f"Datos de {len(probetas)} probetas rellenados")
    
    def _aplicar_formato_fila(self, worksheet, fila: int) -> None:
        """Aplicar formato a una fila de datos"""
        try:
            # Aplicar bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                try:
                    cell = worksheet[f'{col}{fila}']
                    cell.border = thin_border
                    
                    # Aplicar ajuste de texto (text wrapping)
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center',
                        wrap_text=True
                    )
                except Exception:
                    pass
            
            # Alineación específica por columna
            # Columna A (ITEM) - centrado
            try:
                worksheet[f'A{fila}'].alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except Exception:
                pass
            
            # Columna B (ORDEN TRABAJO) - centrado
            try:
                worksheet[f'B{fila}'].alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except Exception:
                pass
            
            # Columna C (CÓDIGO MUESTRA) - centrado
            try:
                worksheet[f'C{fila}'].alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except Exception:
                pass
            
            # Columna D (CLIENTES) - centrado
            try:
                worksheet[f'D{fila}'].alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except Exception:
                pass
            
            # Columna E (FECHA ROTURA) - centrado
            try:
                worksheet[f'E{fila}'].alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except Exception:
                pass
            
            # Columna F (ELEMENTO) - centrado
            try:
                worksheet[f'F{fila}'].alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except Exception:
                pass
            
            # Columna G (F'C) - centrado
            try:
                worksheet[f'G{fila}'].alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except Exception:
                pass
            
            # Columna H (STATUS) - centrado
            try:
                worksheet[f'H{fila}'].alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            except Exception:
                pass
            
        except Exception as e:
            logger.error(f"Error aplicando formato a fila {fila}: {e}")
    
    def _aplicar_formato_final(self, worksheet, total_probetas: int) -> None:
        """Aplicar formato final al documento"""
        try:
            # Ajustar anchos de columnas para mejor presentación
            column_widths = {
                'A': 8,   # ITEM - más estrecho
                'B': 15,  # ORDEN TRABAJO
                'C': 18,  # CÓDIGO MUESTRA LEM
                'D': 25,  # CLIENTES - más ancho para nombres largos
                'E': 18,  # FECHA ROTURA PROGRAMADO
                'F': 15,  # ELEMENTO
                'G': 12,  # F'C
                'H': 15   # STATUS ENSAYADO
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Aplicar formato al header de la tabla (fila 7)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True, size=10)
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                try:
                    cell = worksheet[f'{col}7']
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center',
                        wrap_text=True
                    )
                except Exception:
                    pass
            
            # Ajustar altura de filas para mejor presentación
            for fila in range(7, 7 + total_probetas + 1):
                worksheet.row_dimensions[fila].height = 25.0  # Altura fija para mejor presentación
            
            logger.info("Formato final aplicado al documento")
            
        except Exception as e:
            logger.error(f"Error aplicando formato final: {e}")
    
    def _guardar_archivo(self, workbook, nombre_base: str) -> str:
        """Guardar el archivo Excel generado"""
        import datetime
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{nombre_base}_{timestamp}.xlsx"
        
        # Crear directorio de salida si no existe
        output_dir = os.path.join(os.path.dirname(__file__), '..', 'output')
        os.makedirs(output_dir, exist_ok=True)
        
        output_path = os.path.join(output_dir, filename)
        workbook.save(output_path)
        
        return output_path
    
    def buscar_datos_cliente(self, codigo_muestra: str) -> Optional[Dict[str, Any]]:
        """
        Búsqueda inteligente de datos del cliente basada en el código de muestra
        
        Args:
            codigo_muestra: Código de la muestra (ej: "4259-CO-25")
        
        Returns:
            Diccionario con datos del cliente encontrados o None
        """
        try:
            # Aquí implementarías la lógica de búsqueda en base de datos
            # Por ahora, retorno datos de ejemplo basados en el patrón del código
            
            if not codigo_muestra or '-' not in codigo_muestra:
                return None
            
            partes = codigo_muestra.split('-')
            if len(partes) >= 3:
                # Extraer orden de trabajo del código
                orden_trabajo = f"{partes[0]}-{partes[1]}-{partes[2]}"
                
                # Datos de ejemplo (en producción vendrían de la base de datos)
                datos_cliente = {
                    'orden_trabajo': orden_trabajo,
                    'codigo_documento': 'F-LEM-P-01.09',
                    'version': '04',
                    'fecha_documento': '9/09/2024',
                    'pagina': '1 de 1',
                    'nota': 'SGCMM GROUP',
                    'status_ensayado': 'ROTURADO'
                }
                
                logger.info(f"Datos del cliente encontrados para código: {codigo_muestra}")
                return datos_cliente
            
            return None
            
        except Exception as e:
            logger.error(f"Error en búsqueda de datos del cliente: {e}")
            return None
    
    def buscar_por_codigo_recepcion(self, codigo_recepcion: str) -> Optional[Dict[str, Any]]:
        """
        Búsqueda inteligente de datos del cliente basada en el código de recepción
        
        Args:
            codigo_recepcion: Código de recepción (ej: "1384-25")
        
        Returns:
            Diccionario con datos del cliente encontrados o None
        """
        try:
            # Aquí implementarías la lógica de búsqueda en base de datos por código de recepción
            # Por ahora, retorno datos de ejemplo basados en el patrón del código
            
            if not codigo_recepcion or '-' not in codigo_recepcion:
                return None
            
            # Simular búsqueda por código de recepción
            # En producción, esto buscaría en la tabla de recepciones
            if codigo_recepcion.startswith('1384') or codigo_recepcion.startswith('REC'):
                # Datos de ejemplo para código de recepción
                datos_cliente = {
                    'orden_trabajo': 'OT-1422-25',
                    'codigo_documento': 'F-LEM-P-01.09',
                    'version': '04',
                    'fecha_documento': '9/09/2024',
                    'pagina': '1 de 1',
                    'nota': 'SGCMM GROUP',
                    'status_ensayado': 'ROTURADO',
                    'cliente': 'MEGA CONSTRUCCIONES',
                    'proyecto': 'EDIFICIO RESIDENCIAL',
                    'ubicacion': 'LIMA, PERU'
                }
                
                logger.info(f"Datos del cliente encontrados para código de recepción: {codigo_recepcion}")
                return datos_cliente
            
            return None
            
        except Exception as e:
            logger.error(f"Error en búsqueda por código de recepción: {e}")
            return None
