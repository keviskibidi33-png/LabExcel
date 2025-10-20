"""
Servicio para procesamiento de archivos Excel
Maneja importación, exportación y generación de plantillas
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import tempfile
import os
from datetime import datetime
from typing import List, Dict, Any
from sqlalchemy.orm import Session

from models import RecepcionMuestra, MuestraConcreto
from schemas import RecepcionMuestraCreate, MuestraConcretoCreate

class ExcelService:
    def __init__(self):
        self.template_path = "templates/orden_trabajo_template.xlsx"
    
    def procesar_archivo_excel(self, contenido: bytes, db: Session) -> RecepcionMuestra:
        """Procesar archivo Excel subido y crear orden de trabajo"""
        try:
            # Leer archivo Excel
            df = pd.read_excel(BytesIO(contenido), sheet_name=0, header=None)
            
            # Extraer datos de la orden
            orden_data = self._extraer_datos_orden(df)
            
            # Extraer items
            items_data = self._extraer_items(df)
            
            # Crear orden en base de datos
            recepcion = RecepcionMuestra(**orden_data)
            db.add(recepcion)
            db.flush()  # Para obtener el ID
            
            # Crear items
            for item_data in items_data:
                item = MuestraConcreto(recepcion_id=recepcion.id, **item_data)
                db.add(item)
            
            db.commit()
            db.refresh(recepcion)
            
            return recepcion
            
        except Exception as e:
            db.rollback()
            raise Exception(f"Error procesando archivo Excel: {str(e)}")
    
    def _extraer_datos_orden(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Extraer datos de la orden desde el DataFrame"""
        datos = {}
        
        # Buscar número OT (fila 5, columna 0)
        if len(df) > 5 and pd.notna(df.iloc[5, 0]):
            datos['numero_ot'] = str(df.iloc[5, 0]).strip()
        
        # Buscar número recepción (fila 5, columna 5)
        if len(df) > 5 and pd.notna(df.iloc[5, 5]):
            datos['numero_recepcion'] = str(df.iloc[5, 5]).strip()
        
        # Buscar referencia (fila 5, columna 7)
        if len(df) > 5 and pd.notna(df.iloc[5, 7]):
            datos['referencia'] = str(df.iloc[5, 7]).strip()
        
        # Buscar fechas
        for i in range(len(df)):
            if pd.notna(df.iloc[i, 0]) and "FECHA DE RECEPCIÓN" in str(df.iloc[i, 0]):
                if pd.notna(df.iloc[i, 2]):
                    try:
                        datos['fecha_recepcion'] = pd.to_datetime(df.iloc[i, 2])
                    except:
                        pass
                break
        
        # Buscar plazo de entrega
        for i in range(len(df)):
            if pd.notna(df.iloc[i, 0]) and "PLAZO DE ENTREGA" in str(df.iloc[i, 0]):
                if pd.notna(df.iloc[i, 2]):
                    try:
                        datos['plazo_entrega_dias'] = int(df.iloc[i, 2])
                    except:
                        pass
                break
        
        # Buscar observaciones
        for i in range(len(df)):
            if pd.notna(df.iloc[i, 0]) and "OBSERVACIONES" in str(df.iloc[i, 0]):
                if pd.notna(df.iloc[i, 1]):
                    datos['observaciones'] = str(df.iloc[i, 1]).strip()
                break
        
        # Buscar aperturada por
        for i in range(len(df)):
            if pd.notna(df.iloc[i, 0]) and "O/T APERTURADA POR" in str(df.iloc[i, 0]):
                if pd.notna(df.iloc[i, 2]):
                    datos['aperturada_por'] = str(df.iloc[i, 2]).strip()
                break
        
        # Buscar designada a
        for i in range(len(df)):
            if pd.notna(df.iloc[i, 0]) and "OT DESIGADA A" in str(df.iloc[i, 0]):
                if pd.notna(df.iloc[i, 2]):
                    datos['designada_a'] = str(df.iloc[i, 2]).strip()
                break
        
        return datos
    
    def _extraer_items(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Extraer items de la orden desde el DataFrame"""
        items = []
        
        # Buscar la sección de items (desde fila 8 hasta encontrar fechas)
        # Ajustar el rango para encontrar datos más abajo en el Excel
        for i in range(8, len(df)):
            if pd.notna(df.iloc[i, 0]) and "FECHA DE RECEPCIÓN" in str(df.iloc[i, 0]):
                break
            
            # Verificar si hay datos en las columnas de item
            # Solo requerimos que la columna A (N°) tenga un número válido
            if (pd.notna(df.iloc[i, 0]) and str(df.iloc[i, 0]).strip().isdigit()):
                
                # Manejar tipos de datos correctamente
                cantidad = 1
                if pd.notna(df.iloc[i, 7]):
                    try:
                        cantidad = int(df.iloc[i, 7])
                    except (ValueError, TypeError):
                        cantidad = 1
                
                item_data = {
                    'item_numero': int(df.iloc[i, 0]),
                    'codigo_muestra_lem': str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else "",  # Columna B: Código muestra LEM
                    'codigo_muestra': str(df.iloc[i, 2]).strip() if pd.notna(df.iloc[i, 2]) else "",  # Columna C: Codigo
                    'estructura': str(df.iloc[i, 3]).strip() if pd.notna(df.iloc[i, 3]) else "",       # Columna D: Estructura
                    'cantidad': cantidad,
                    'especificacion': str(df.iloc[i, 8]).strip() if pd.notna(df.iloc[i, 8]) else ""
                }
                items.append(item_data)
        
        return items
    
    def generar_plantilla_excel(self, recepcion: RecepcionMuestra = None) -> str:
        """Generar plantilla Excel basada en el diseño del PDF"""
        from templates.recepcion_muestra_template import RecepcionMuestraTemplate
        
        template = RecepcionMuestraTemplate()
        wb = template.crear_plantilla_vacia()
        
        # Si se proporciona una orden, prellenar los datos
        if recepcion:
            self._prellenar_datos_recepcion(wb, recepcion)
        
        # Guardar archivo temporal
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def _prellenar_datos_recepcion(self, wb, recepcion: RecepcionMuestra):
        """Prellenar la plantilla de recepción con datos de una orden existente"""
        ws = wb.active
        
        # Datos principales
        if recepcion.numero_recepcion:
            ws['C8'] = recepcion.numero_recepcion
        if recepcion.fecha_recepcion:
            ws['H8'] = recepcion.fecha_recepcion.strftime("%d/%m/%Y")
        if recepcion.numero_ot:
            ws['H9'] = recepcion.numero_ot
        
        # Llenar muestras
        def safe_set_cell(cell_ref, value):
            cell = ws[cell_ref]
            target_cell = cell
            # Resolver celdas fusionadas si aplica
            for merged_range in ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    top_left = merged_range.min_row, merged_range.min_col
                    target_cell = ws.cell(row=top_left[0], column=top_left[1])
                    break
            target_cell.value = value
            return target_cell

        row = 23  # Unificar con servicios colaborativos: la tabla inicia en fila 23
        for item in recepcion.muestras:
            if row <= 43:  # Máximo 21 filas (23..43)
                # Limpiar celdas de la fila antes de escribir
                for col in ['A','B','C','D','E','F','G','H','I','J','K']:
                    try:
                        ws[f'{col}{row}'].value = None
                    except Exception:
                        pass

                # Secuencia: N° (A) → Código muestra LEM (B) → Identificación (D) → Estructura (E)
                # A: número secuencial
                safe_set_cell(f'A{row}', getattr(item, 'item_numero', None) or (row - 22))

                # B: código muestra LEM (forzar formato texto para no perder ceros ni notación)
                b_cell = safe_set_cell(f'B{row}', getattr(item, 'codigo_muestra_lem', ''))
                try:
                    b_cell.number_format = '@'
                except Exception:
                    pass

                # D/E y demás campos
                safe_set_cell(f'D{row}', getattr(item, 'identificacion_muestra', ''))
                safe_set_cell(f'E{row}', getattr(item, 'estructura', ''))
                safe_set_cell(f'F{row}', getattr(item, 'fc_kg_cm2', 280))
                if hasattr(item, 'fecha_moldeo') and item.fecha_moldeo:
                    safe_set_cell(f'G{row}', item.fecha_moldeo.strftime("%d/%m/%y"))
                if hasattr(item, 'hora_moldeo') and item.hora_moldeo:
                    safe_set_cell(f'H{row}', item.hora_moldeo)
                safe_set_cell(f'I{row}', getattr(item, 'edad', 10))
                if hasattr(item, 'fecha_rotura') and item.fecha_rotura:
                    safe_set_cell(f'J{row}', item.fecha_rotura.strftime("%d/%m/%y"))
                safe_set_cell(f'K{row}', 'SI' if getattr(item, 'requiere_densidad', False) else 'NO')
                row += 1
        
        # Fecha estimada de culminación
        if recepcion.fecha_fin_programado:
            ws['F47'] = recepcion.fecha_fin_programado.strftime("%d/%m/%Y")
        
        # Responsables
        if recepcion.aperturada_por:
            ws['D50'] = recepcion.aperturada_por
        if recepcion.designada_a:
            ws['D51'] = recepcion.designada_a
    
    def exportar_ordenes(self, orden_ids: List[int], db: Session) -> str:
        """Exportar múltiples órdenes a un archivo Excel"""
        recepciones = db.query(RecepcionMuestra).filter(RecepcionMuestra.id.in_(orden_ids)).all()
        
        if not recepciones:
            raise Exception("No se encontraron órdenes para exportar")
        
        wb = openpyxl.Workbook()
        
        # Crear hoja de resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Órdenes"
        
        # Encabezados del resumen
        headers = ["ID", "N° OT", "N° RECEPCIÓN", "REFERENCIA", "FECHA CREACIÓN", 
                  "ESTADO", "ITEMS", "APERTURADA POR", "DESIGNADA A"]
        
        for col, header in enumerate(headers, 1):
            ws_resumen.cell(row=1, column=col, value=header)
            ws_resumen.cell(row=1, column=col).font = Font(bold=True)
        
        # Llenar datos del resumen
        for row, recepcion in enumerate(recepciones, 2):
            ws_resumen.cell(row=row, column=1, value=recepcion.id)
            ws_resumen.cell(row=row, column=2, value=recepcion.numero_ot)
            ws_resumen.cell(row=row, column=3, value=recepcion.numero_recepcion)
            ws_resumen.cell(row=row, column=4, value=recepcion.codigo_trazabilidad or "")
            ws_resumen.cell(row=row, column=5, value=recepcion.fecha_creacion.strftime("%d/%m/%Y"))
            ws_resumen.cell(row=row, column=6, value=recepcion.estado)
            ws_resumen.cell(row=row, column=7, value=len(recepcion.muestras))
            ws_resumen.cell(row=row, column=8, value=recepcion.aperturada_por or "")
            ws_resumen.cell(row=row, column=9, value=recepcion.designada_a or "")
        
        # Crear hoja para cada recepción
        for recepcion in recepciones:
            ws = wb.create_sheet(title=f"OT-{recepcion.numero_ot}")
            self._llenar_hoja_orden(ws, recepcion)
        
        # Guardar archivo temporal
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def _llenar_hoja_orden(self, ws, recepcion: RecepcionMuestra):
        """Llenar una hoja con los datos de una orden específica"""
        # Implementar llenado de hoja individual
        # Similar a generar_plantilla_excel pero para exportación
        pass
