import io
from copy import copy
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font

class OTExcelCollaborativeService:
    """Servicio para generar archivos Excel de Órdenes de Trabajo"""
    
    # Constantes para el template
    TEMPLATE_PATH = "templates/orden_trabajo_template.xlsx"
    FILA_INICIO_ITEMS = 10
    
    def __init__(self):
        self.template_path = self.TEMPLATE_PATH
    
    def modificar_excel_con_datos(self, ot_data: Dict[str, Any], items: List[Dict[str, Any]], 
                                 template_path: Optional[str] = None) -> bytes:
        """Modificar archivo Excel existente con datos del formulario"""
        
        # Ruta del template
        template_file = template_path or self.template_path
        
        # Cargar template directamente
        workbook = openpyxl.load_workbook(template_file)
        worksheet = workbook.active
        
        # Rellenar datos
        self._rellenar_datos_ot(worksheet, ot_data)
        self._rellenar_datos_items(worksheet, items)
        
        # Ajustar ancho de columna C para mejor visualización
        self._ajustar_ancho_columnas(worksheet)
        
        # Guardar en memoria
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    def _rellenar_datos_ot(self, worksheet, ot_data: Dict[str, Any]):
        """Rellenar datos de la orden de trabajo en el Excel"""
        
        def safe_set_cell(cell_ref, value, center_align=False):
            try:
                if isinstance(value, str):
                    value = value.strip()
                if value is None or value == '':
                    return
                
                # Escribir directamente en la celda
                worksheet[cell_ref].value = value
                
                # Aplicar alineación centrada si se solicita
                if center_align:
                    worksheet[cell_ref].alignment = Alignment(horizontal='center', vertical='center')
                
            except Exception as e:
                print(f"Error escribiendo en {cell_ref}: {e}")
                pass
        
        def safe_merge_and_set_cell(start_cell, end_cell, value):
            """Fusionar celdas y escribir valor"""
            try:
                if isinstance(value, str):
                    value = value.strip()
                if value is None or value == '':
                    return
                
                # Fusionar celdas
                worksheet.merge_cells(f'{start_cell}:{end_cell}')
                
                # Escribir valor en la celda superior izquierda
                worksheet[start_cell].value = value
                
                # Aplicar alineación (centrada)
                worksheet[start_cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Ajustar altura de fila basada en el contenido (sistema adaptativo)
                fila_numero = int(start_cell[1:])  # Extraer número de fila
                altura_base = 15
                
                # Sistema simplificado: solo 2 casos
                longitud_texto = len(value)
                
                # Contar párrafos (líneas separadas por punto y seguido)
                parrafos = value.count('. ') + 1  # Aproximación de párrafos
                
                if parrafos <= 1 and longitud_texto <= 100:
                    # 1 párrafo o menos: usar altura del template (no modificar)
                    altura_final = None  # No modificar altura
                else:
                    # Más de 1 párrafo: altura reducida para descripción
                    altura_adicional = max(15, longitud_texto // 60) * 4  # Más pequeño
                    altura_final = altura_base + altura_adicional
                
                # Solo modificar altura si es necesario
                if altura_final is not None:
                    worksheet.row_dimensions[fila_numero].height = altura_final
                    print(f"Fusionado {start_cell}:{end_cell} con valor: '{value}' (altura: {altura_final})")
                else:
                    print(f"Fusionado {start_cell}:{end_cell} con valor: '{value}' (altura: template original)")
                
            except Exception as e:
                print(f"Error fusionando celdas {start_cell}:{end_cell}: {e}")
                pass
        
        def safe_merge_observaciones(start_cell, end_cell, value):
            """Fusionar celdas para observaciones con altura original del template"""
            try:
                if isinstance(value, str):
                    value = value.strip()
                if value is None or value == '':
                    return
                
                # Fusionar celdas
                worksheet.merge_cells(f'{start_cell}:{end_cell}')
                
                # Escribir valor con prefijo "OBSERVACIONES:" en la celda superior izquierda
                worksheet[start_cell].value = f"OBSERVACIONES: {value}"
                
                # Aplicar alineación (izquierda, superior) para observaciones
                worksheet[start_cell].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # NO modificar altura - usar la altura original del template
                print(f"Fusionado {start_cell}:{end_cell} con valor: '{value}' (altura: template original)")
                
            except Exception as e:
                print(f"Error fusionando observaciones {start_cell}:{end_cell}: {e}")
                pass
        
        # Datos principales - CENTRADOS
        safe_set_cell('B6', ot_data.get('numero_ot', ''), center_align=True)
        safe_set_cell('F6', ot_data.get('numero_recepcion', ''), center_align=True)
        
        # Fechas y plazos - CENTRADOS
        safe_set_cell('C31', ot_data.get('fecha_recepcion', ''), center_align=True)
        safe_set_cell('C32', ot_data.get('plazo_entrega_dias', ''), center_align=True)
        
        # Fechas programadas - CENTRADOS
        safe_set_cell('E31', ot_data.get('fecha_inicio_programado', ''), center_align=True)
        safe_set_cell('E32', ot_data.get('fecha_fin_programado', ''), center_align=True)
        
        # Fechas reales - CENTRADOS
        safe_set_cell('G31', ot_data.get('fecha_inicio_real', ''), center_align=True)
        safe_set_cell('G32', ot_data.get('fecha_fin_real', ''), center_align=True)
        
        # Variaciones - CENTRADOS
        safe_set_cell('I31', ot_data.get('variacion_inicio', ''), center_align=True)
        safe_set_cell('I32', ot_data.get('variacion_fin', ''), center_align=True)
        
        # Duración real - CENTRADA (corregida a D33)
        safe_set_cell('D33', ot_data.get('duracion_real_dias', ''), center_align=True)
        
        # Observaciones fusionadas desde A34 hasta I37 para evitar que se escape el texto
        safe_merge_observaciones('A34', 'I37', ot_data.get('observaciones', ''))
        
        # Responsables - CENTRADOS
        safe_set_cell('D38', ot_data.get('aperturada_por', ''), center_align=True)  # Cambiado a D38
        safe_set_cell('H38', ot_data.get('designada_a', ''), center_align=True)
    
    def _rellenar_datos_items(self, worksheet, items: List[Dict[str, Any]]):
        """Rellenar datos de los items en la tabla"""
        
        def safe_set_cell(cell_ref, value, center_align=False):
            try:
                if isinstance(value, str):
                    value = value.strip()
                if value is None or value == '':
                    return
                
                worksheet[cell_ref].value = value
                
                # Aplicar alineación centrada si se solicita
                if center_align:
                    worksheet[cell_ref].alignment = Alignment(horizontal='center', vertical='center')
                
            except Exception as e:
                print(f"Error escribiendo item en {cell_ref}: {e}")
                pass
        
        def safe_merge_and_set_cell(start_cell, end_cell, value):
            """Fusionar celdas y escribir valor"""
            try:
                if isinstance(value, str):
                    value = value.strip()
                if value is None or value == '':
                    return
                
                # Fusionar celdas
                worksheet.merge_cells(f'{start_cell}:{end_cell}')
                
                # Escribir valor en la celda superior izquierda
                worksheet[start_cell].value = value
                
                # Aplicar alineación (centrada)
                worksheet[start_cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Ajustar altura de fila basada en el contenido (sistema adaptativo)
                fila_numero = int(start_cell[1:])  # Extraer número de fila
                altura_base = 15
                
                # Sistema simplificado: solo 2 casos
                longitud_texto = len(value)
                
                # Contar párrafos (líneas separadas por punto y seguido)
                parrafos = value.count('. ') + 1  # Aproximación de párrafos
                
                if parrafos <= 1 and longitud_texto <= 100:
                    # 1 párrafo o menos: usar altura del template (no modificar)
                    altura_final = None  # No modificar altura
                else:
                    # Más de 1 párrafo: altura reducida para descripción
                    altura_adicional = max(15, longitud_texto // 60) * 4  # Más pequeño
                    altura_final = altura_base + altura_adicional
                
                # Solo modificar altura si es necesario
                if altura_final is not None:
                    worksheet.row_dimensions[fila_numero].height = altura_final
                    print(f"Fusionado {start_cell}:{end_cell} con valor: '{value}' (altura: {altura_final})")
                else:
                    print(f"Fusionado {start_cell}:{end_cell} con valor: '{value}' (altura: template original)")
                
            except Exception as e:
                print(f"Error fusionando celdas {start_cell}:{end_cell}: {e}")
                pass
        
        # Columnas: A=ÍTEM, B=CÓDIGO, D-H=DESCRIPCIÓN (fusionada), I=CANTIDAD
        fila_inicio = 10
        
        for indice, item in enumerate(items):
            fila_actual = fila_inicio + indice
            
            # ÍTEM centrado
            safe_set_cell(f'A{fila_actual}', item.get('item_numero') or indice + 1, center_align=True)
            
            # CÓDIGO DE MUESTRA fusionado en B:C con altura adaptativa
            safe_merge_and_set_cell(f'B{fila_actual}', f'C{fila_actual}', item.get('codigo_muestra', ''))
            
            # DESCRIPCIÓN en celdas fusionadas D:H (5 columnas) para más espacio
            safe_merge_and_set_cell(f'D{fila_actual}', f'H{fila_actual}', item.get('descripcion', ''))
            
            safe_set_cell(f'I{fila_actual}', item.get('cantidad', ''), center_align=True)
    
    def _ajustar_ancho_columnas(self, worksheet):
        """Ajustar ancho de columnas para mejor visualización"""
        try:
            # Hacer la columna C más ancha para que los datos no se vean "juntos"
            worksheet.column_dimensions['C'].width = 25  # Ancho más largo para mejor visualización
            print("Ancho de columna C ajustado a 25")
        except Exception as e:
            print(f"Error ajustando ancho de columnas: {e}")
            pass