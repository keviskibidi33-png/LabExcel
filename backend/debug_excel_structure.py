#!/usr/bin/env python3
"""
Debug del archivo Excel generado para identificar problemas de estructura
"""

import sys
import os
import openpyxl
from openpyxl import load_workbook

# Agregar el directorio backend al path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def debug_excel_structure():
    """Analizar la estructura del Excel generado"""
    
    try:
        print("=== DIAGNOSTICO DE ESTRUCTURA EXCEL ===")
        
        # Cargar el template original
        template_path = "templates/recepcion_template.xlsx"
        print(f"1. Cargando template: {template_path}")
        
        workbook = load_workbook(template_path)
        worksheet = workbook.active
        
        print(f"2. Template cargado exitosamente")
        print(f"   - Filas: {worksheet.max_row}")
        print(f"   - Columnas: {worksheet.max_column}")
        
        # Analizar celdas fusionadas
        print(f"3. Analizando celdas fusionadas...")
        merged_ranges = list(worksheet.merged_cells.ranges)
        print(f"   - Total de rangos fusionados: {len(merged_ranges)}")
        
        # Verificar rangos fusionados problemáticos
        problematic_ranges = []
        for merged_range in merged_ranges:
            # Verificar si el rango es muy ancho (más de 5 columnas)
            width = merged_range.max_col - merged_range.min_col + 1
            height = merged_range.max_row - merged_range.min_row + 1
            
            if width > 5 or height > 3:
                problematic_ranges.append({
                    'range': str(merged_range),
                    'width': width,
                    'height': height
                })
        
        if problematic_ranges:
            print(f"   - Rangos fusionados problemáticos encontrados:")
            for pr in problematic_ranges:
                print(f"     * {pr['range']} (ancho: {pr['width']}, alto: {pr['height']})")
        else:
            print(f"   - No se encontraron rangos fusionados problemáticos")
        
        # Verificar celdas vacías o con valores extraños
        print(f"4. Verificando contenido de celdas...")
        empty_cells = 0
        error_cells = 0
        
        for row in range(1, min(worksheet.max_row + 1, 100)):  # Solo primeras 100 filas
            for col in range(1, min(worksheet.max_column + 1, 20)):  # Solo primeras 20 columnas
                try:
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is None:
                        empty_cells += 1
                    elif isinstance(cell.value, str) and len(cell.value) > 1000:
                        error_cells += 1
                        print(f"     * Celda {cell.coordinate} tiene contenido muy largo: {len(cell.value)} caracteres")
                except Exception as e:
                    error_cells += 1
                    print(f"     * Error en celda {row},{col}: {e}")
        
        print(f"   - Celdas vacías: {empty_cells}")
        print(f"   - Celdas con errores: {error_cells}")
        
        # Verificar estilos
        print(f"5. Verificando estilos...")
        style_errors = 0
        for row in range(1, min(worksheet.max_row + 1, 50)):
            for col in range(1, min(worksheet.max_column + 1, 15)):
                try:
                    cell = worksheet.cell(row=row, column=col)
                    # Verificar si el estilo es válido
                    _ = cell.border
                    _ = cell.fill
                    _ = cell.font
                    _ = cell.alignment
                except Exception as e:
                    style_errors += 1
                    if style_errors <= 5:  # Solo mostrar primeros 5 errores
                        print(f"     * Error de estilo en {cell.coordinate}: {e}")
        
        print(f"   - Errores de estilo: {style_errors}")
        
        # Verificar dimensiones de filas y columnas
        print(f"6. Verificando dimensiones...")
        row_errors = 0
        col_errors = 0
        
        for row in range(1, min(worksheet.max_row + 1, 100)):
            try:
                height = worksheet.row_dimensions[row].height
                if height and (height < 0 or height > 1000):
                    row_errors += 1
            except Exception as e:
                row_errors += 1
        
        for col in range(1, min(worksheet.max_column + 1, 20)):
            try:
                width = worksheet.column_dimensions[chr(64 + col)].width
                if width and (width < 0 or width > 1000):
                    col_errors += 1
            except Exception as e:
                col_errors += 1
        
        print(f"   - Errores de altura de fila: {row_errors}")
        print(f"   - Errores de ancho de columna: {col_errors}")
        
        # Resumen
        print(f"\n=== RESUMEN ===")
        total_issues = len(problematic_ranges) + error_cells + style_errors + row_errors + col_errors
        if total_issues == 0:
            print("Template Excel esta en buen estado")
        else:
            print(f"Se encontraron {total_issues} problemas potenciales")
            print("   - Esto puede causar el error de Excel al abrir el archivo")
        
        return total_issues == 0
        
    except Exception as e:
        print(f"Error analizando Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = debug_excel_structure()
    if success:
        print("\nDIAGNOSTICO EXITOSO")
    else:
        print("\nDIAGNOSTICO FALLIDO")
