"""
Generador de plantilla Excel para RECEPCIÓN DE MUESTRA CILÍNDRICAS DE CONCRETO
Replica exactamente el diseño del PDF original
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import tempfile
import os

class RecepcionMuestraTemplate:
    def __init__(self):
        self.wb = None
        self.ws = None
        
    def crear_plantilla_vacia(self):
        """Crear plantilla Excel vacía basada en el diseño exacto del PDF"""
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Recepcion Muestra"
        
        # Configurar estilos
        self._configurar_estilos()
        
        # Crear estructura del formulario exacta
        self._crear_encabezado()
        self._crear_datos_principales()
        self._crear_datos_cliente()
        self._crear_datos_solicitante()
        self._crear_tabla_muestras()
        self._crear_emision_informes()
        self._crear_responsables()
        self._crear_pie_pagina()
        
        # Ajustar formato
        self._ajustar_formato()
        
        return self.wb
    
    def _configurar_estilos(self):
        """Configurar estilos para el documento"""
        self.estilos = {
            'titulo_principal': Font(name='Arial', size=12, bold=True),
            'titulo_seccion': Font(name='Arial', size=10, bold=True),
            'texto_normal': Font(name='Arial', size=9),
            'texto_pequeno': Font(name='Arial', size=8),
            'borde_completo': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'borde_superior': Border(top=Side(style='thin')),
            'borde_inferior': Border(bottom=Side(style='thin')),
            'centrado': Alignment(horizontal='center', vertical='center'),
            'izquierda': Alignment(horizontal='left', vertical='center'),
            'relleno_gris': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        }
    
    def _crear_encabezado(self):
        """Crear encabezado del documento"""
        # Título principal (celda A1)
        self.ws['A1'] = "RECEPCION DE MUESTRA CILINDRICAS DE"
        self.ws['A1'].font = self.estilos['titulo_principal']
        self.ws['A1'].alignment = self.estilos['centrado']
        self.ws.merge_cells('A1:J1')
        
        # Segunda línea del título (celda A2)
        self.ws['A2'] = "CONCRETO"
        self.ws['A2'].font = self.estilos['titulo_principal']
        self.ws['A2'].alignment = self.estilos['centrado']
        self.ws.merge_cells('A2:J2')
        
        # Información del documento (celdas H3-J6)
        self.ws['H3'] = "CÓDIGO:"
        self.ws['I3'] = "F-LEM-P-01.02"
        self.ws['H4'] = "VERSIÓN:"
        self.ws['I4'] = "07"
        self.ws['H5'] = "FECHA:"
        self.ws['I5'] = datetime.now().strftime("%d/%m/%Y")
        self.ws['H6'] = "PÁGINA:"
        self.ws['I6'] = "1 de 1"
        
        # Aplicar estilos
        for cell in ['H3', 'H4', 'H5', 'H6']:
            self.ws[cell].font = self.estilos['texto_normal']
        for cell in ['I3', 'I4', 'I5', 'I6']:
            self.ws[cell].font = self.estilos['texto_normal']
    
    def _crear_datos_principales(self):
        """Crear sección de datos principales"""
        # Línea de separación (fila 7)
        self.ws['A7'] = ""
        
        # Datos principales (fila 8)
        self.ws['A8'] = "RECEPCIÓN N°:"
        self.ws['C8'] = ""  # Campo editable
        self.ws['F8'] = "FECHA DE RECEPCIÓN:"
        self.ws['H8'] = ""  # Campo editable
        
        # Datos principales (fila 9)
        self.ws['A9'] = "COTIZACIÓN N°:"
        self.ws['C9'] = ""  # Campo editable
        self.ws['F9'] = "OT N°:"
        self.ws['H9'] = ""  # Campo editable
        
        # Nota sobre código de trazabilidad (fila 10)
        self.ws['C10'] = "(Código de trazabilidad con el informe de ensayo)"
        self.ws['C10'].font = self.estilos['texto_pequeno']
        self.ws['C10'].alignment = self.estilos['izquierda']
        
        # Asunto (fila 11)
        self.ws['A11'] = "ASUNTO: SOLICITO EJECUCIÓN DE ENSAYOS"
        self.ws['A11'].font = self.estilos['texto_normal']
        self.ws.merge_cells('A11:J11')
        
        # Aplicar estilos
        for cell in ['A8', 'F8', 'A9', 'F9']:
            self.ws[cell].font = self.estilos['texto_normal']
        for cell in ['C8', 'H8', 'C9', 'H9']:
            self.ws[cell].font = self.estilos['texto_normal']
            self.ws[cell].border = self.estilos['borde_inferior']
    
    def _crear_datos_cliente(self):
        """Crear sección de datos del cliente"""
        # Título de sección (fila 12)
        self.ws['A12'] = "Cliente:"
        self.ws['C12'] = ""  # Campo editable
        self.ws['F12'] = "DATOS PARA FACTURACION Y PERSONA DE CONTACTO PARA EL ENVIO DEL INFORME DE LABORATORIO"
        self.ws['F12'].font = self.estilos['titulo_seccion']
        self.ws.merge_cells('F12:J12')
        
        # Domicilio legal (fila 13)
        self.ws['A13'] = "Domicilio legal:"
        self.ws['C13'] = ""  # Campo editable
        self.ws.merge_cells('C13:J13')
        
        # RUC (fila 14)
        self.ws['A14'] = "RUC:"
        self.ws['C14'] = ""  # Campo editable
        
        # Persona contacto (fila 15)
        self.ws['A15'] = "Persona contacto:"
        self.ws['C15'] = ""  # Campo editable
        self.ws['F15'] = "Teléfono:"
        self.ws['H15'] = ""  # Campo editable
        
        # Email (fila 16)
        self.ws['A16'] = "E-MAIL:"
        self.ws['C16'] = ""  # Campo editable
        self.ws.merge_cells('C16:J16')
        
        # Aplicar estilos
        for cell in ['A12', 'A13', 'A14', 'A15', 'A16', 'F15']:
            self.ws[cell].font = self.estilos['texto_normal']
        for cell in ['C12', 'C13', 'C14', 'C15', 'C16', 'H15']:
            self.ws[cell].font = self.estilos['texto_normal']
            self.ws[cell].border = self.estilos['borde_inferior']
    
    def _crear_datos_solicitante(self):
        """Crear sección de datos del solicitante"""
        # Título de sección (fila 17)
        self.ws['A17'] = "DATOS A CONSIGNAR EN EL INFORME DE ENSAYO"
        self.ws['A17'].font = self.estilos['titulo_seccion']
        self.ws.merge_cells('A17:J17')
        
        # Solicitante (fila 18)
        self.ws['A18'] = "Solicitante:"
        self.ws['C18'] = ""  # Campo editable
        self.ws['F18'] = "DATOS QUE IRA EN EL INFORME DE LABORATORIO"
        self.ws['F18'].font = self.estilos['titulo_seccion']
        self.ws.merge_cells('F18:J18')
        
        # Domicilio legal solicitante (fila 19)
        self.ws['A19'] = "Domicilio legal:"
        self.ws['C19'] = ""  # Campo editable
        self.ws.merge_cells('C19:J19')
        
        # Proyecto (fila 20)
        self.ws['A20'] = "Proyecto:"
        self.ws['C20'] = ""  # Campo editable
        self.ws.merge_cells('C20:J20')
        
        # Ubicación (fila 21)
        self.ws['A21'] = "Ubicación:"
        self.ws['C21'] = ""  # Campo editable
        self.ws.merge_cells('C21:J21')
        
        # Nota (fila 22)
        self.ws['A22'] = "* No llenar la zona sombreada"
        self.ws['A22'].font = self.estilos['texto_pequeno']
        self.ws.merge_cells('A22:J22')
        
        # Aplicar estilos
        for cell in ['A18', 'A19', 'A20', 'A21']:
            self.ws[cell].font = self.estilos['texto_normal']
        for cell in ['C18', 'C19', 'C20', 'C21']:
            self.ws[cell].font = self.estilos['texto_normal']
            self.ws[cell].border = self.estilos['borde_inferior']
    
    def _crear_tabla_muestras(self):
        """Crear tabla de muestras"""
        # Encabezados de la tabla (fila 23)
        self.ws['A23'] = "N°"
        self.ws['B23'] = "Código muestra LEM"
        self.ws['C23'] = "Identificacion muestra"
        self.ws['D23'] = "Codigo"
        self.ws['E23'] = "Estructura"
        self.ws['F23'] = "F'c (kg/cm2)"
        self.ws['G23'] = "Fecha moldeo"
        self.ws['H23'] = "Hora de moldeo (1)"
        self.ws['I23'] = "Edad"
        self.ws['J23'] = "Fecha rotura"
        
        # Aplicar estilos a encabezados
        for col in range(1, 11):  # A a J
            cell = self.ws.cell(row=23, column=col)
            cell.font = self.estilos['titulo_seccion']
            cell.alignment = self.estilos['centrado']
            cell.border = self.estilos['borde_completo']
            cell.fill = self.estilos['relleno_gris']
        
        # Crear filas para muestras (máximo 20 muestras)
        for i in range(24, 44):  # Filas 24 a 43
            self.ws[f'A{i}'] = i - 23  # Número de muestra
            self.ws[f'A{i}'].font = self.estilos['texto_normal']
            self.ws[f'A{i}'].alignment = self.estilos['centrado']
            self.ws[f'A{i}'].border = self.estilos['borde_completo']
            
            # Campos editables
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                self.ws[f'{col}{i}'].border = self.estilos['borde_completo']
                self.ws[f'{col}{i}'].font = self.estilos['texto_normal']
        
        # Nota sobre hora de moldeo (fila 44)
        self.ws['A44'] = "(1) OBLIGATORIO: para edades menores de 3 días."
        self.ws['A44'].font = self.estilos['texto_pequeno']
        self.ws.merge_cells('A44:J44')
    
    def _crear_emision_informes(self):
        """Crear sección de emisión de informes"""
        # Título (fila 46)
        self.ws['A46'] = "Emisión de Informes:"
        self.ws['A46'].font = self.estilos['texto_normal']
        self.ws['F46'] = "Fecha estimada de culminación de ensayos a partir de la aceptación:"
        self.ws['F46'].font = self.estilos['texto_normal']
        self.ws.merge_cells('F46:J46')
        
        # Opciones de emisión (fila 47)
        self.ws['A47'] = "- Físico (El cliente recoger los informes en el laboratorio)"
        self.ws['A47'].font = self.estilos['texto_normal']
        self.ws['F47'] = ""  # Campo editable para fecha
        self.ws['F47'].border = self.estilos['borde_inferior']
        
        # Opción digital (fila 48)
        self.ws['A48'] = "- Digital (Envío a los correos autorizados, con firma digital)"
        self.ws['A48'].font = self.estilos['texto_normal']
    
    def _crear_responsables(self):
        """Crear sección de responsables"""
        # Entregado por (fila 50)
        self.ws['A50'] = "Entregado por:"
        self.ws['A50'].font = self.estilos['texto_normal']
        self.ws['C50'] = "(Cliente)"
        self.ws['C50'].font = self.estilos['texto_pequeno']
        self.ws['D50'] = ""  # Campo editable
        self.ws['D50'].border = self.estilos['borde_inferior']
        
        # Recibido por (fila 51)
        self.ws['A51'] = "Recepcionado por:"
        self.ws['A51'].font = self.estilos['texto_normal']
        self.ws['C51'] = "(Laboratorio GEOFAL)"
        self.ws['C51'].font = self.estilos['texto_pequeno']
        self.ws['D51'] = ""  # Campo editable
        self.ws['D51'].border = self.estilos['borde_inferior']
    
    def _crear_pie_pagina(self):
        """Crear pie de página"""
        # Información de contacto (fila 53)
        self.ws['A53'] = "Web: www.geofal.com.pe / Correo: laboratorio@geofal.com.pe / Av. Marañón N°763 Los Olivos, Lima / Teléfono: 01-7543070"
        self.ws['A53'].font = self.estilos['texto_pequeno']
        self.ws['A53'].alignment = self.estilos['centrado']
        self.ws.merge_cells('A53:J53')
    
    def _ajustar_formato(self):
        """Ajustar formato general del documento"""
        # Ajustar ancho de columnas
        column_widths = {
            'A': 8,   # N°
            'B': 18,  # Código muestra LEM
            'C': 20,  # Identificación muestra
            'D': 12,  # Código
            'E': 25,  # Estructura
            'F': 12,  # F'c
            'G': 12,  # Fecha moldeo
            'H': 15,  # Hora moldeo
            'I': 8,   # Edad
            'J': 12   # Fecha rotura
        }
        
        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width
        
        # Ajustar altura de filas
        for row in range(1, 60):
            self.ws.row_dimensions[row].height = 15
    
    def guardar_plantilla(self, ruta_archivo=None):
        """Guardar la plantilla en un archivo"""
        if ruta_archivo is None:
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            ruta_archivo = temp_file.name
            temp_file.close()
        
        self.wb.save(ruta_archivo)
        return ruta_archivo

def crear_plantilla_recepcion_muestra():
    """Función principal para crear la plantilla de recepción"""
    template = RecepcionMuestraTemplate()
    wb = template.crear_plantilla_vacia()
    ruta = template.guardar_plantilla()
    return ruta

if __name__ == "__main__":
    # Crear plantilla de ejemplo
    ruta_plantilla = crear_plantilla_recepcion_muestra()
    print(f"Plantilla de recepción creada en: {ruta_plantilla}")
