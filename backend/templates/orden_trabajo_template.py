"""
Generador de template Excel para órdenes de trabajo
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def crear_template_orden_trabajo():
    """Crear template Excel para órdenes de trabajo"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orden de Trabajo"
    
    # Estilos
    header_font = Font(name='Arial', size=12, bold=True)
    title_font = Font(name='Arial', size=14, bold=True)
    normal_font = Font(name='Arial', size=10)
    small_font = Font(name='Arial', size=9)
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezado
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ORDEN DE TRABAJO'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    # Información del laboratorio
    ws['A2'] = 'CÓDIGO: F-LEM-P-01.02'
    ws['A2'].font = small_font
    ws['C2'] = 'VERSIÓN: 07'
    ws['C2'].font = small_font
    ws['E2'] = f'FECHA: {openpyxl.utils.datetime.datetime.now().strftime("%d/%m/%Y")}'
    ws['E2'].font = small_font
    
    # Datos principales
    ws['A4'] = 'N° OT:'
    ws['A4'].font = header_font
    ws['C4'] = 'Nº RECEPCIÓN:'
    ws['C4'].font = header_font
    
    ws['A5'] = 'Cliente:'
    ws['A5'].font = header_font
    ws['C5'] = 'RUC:'
    ws['C5'].font = header_font
    ws['E5'] = 'Teléfono:'
    ws['E5'].font = header_font
    
    ws['A6'] = 'Domicilio Legal:'
    ws['A6'].font = header_font
    
    ws['A7'] = 'Persona Contacto:'
    ws['A7'].font = header_font
    ws['C7'] = 'Email:'
    ws['C7'].font = header_font
    
    # Datos del solicitante
    ws['A9'] = 'Solicitante:'
    ws['A9'].font = header_font
    ws['C9'] = 'Proyecto:'
    ws['C9'].font = header_font
    
    ws['A10'] = 'Domicilio Solicitante:'
    ws['A10'].font = header_font
    
    ws['A11'] = 'Ubicación:'
    ws['A11'].font = header_font
    
    # Fechas y plazos
    ws['A13'] = 'FECHA DE RECEPCIÓN:'
    ws['A13'].font = header_font
    ws['C13'] = 'PLAZO DE ENTREGA (DIAS):'
    ws['C13'].font = header_font
    
    ws['A14'] = 'INICIO PROGRAMADO:'
    ws['A14'].font = header_font
    ws['C14'] = 'FIN PROGRAMADO:'
    ws['C14'].font = header_font
    
    ws['A15'] = 'INICIO REAL:'
    ws['A15'].font = header_font
    ws['C15'] = 'FIN REAL:'
    ws['C15'].font = header_font
    
    ws['A16'] = 'VARIACION DE INICIO:'
    ws['A16'].font = header_font
    ws['C16'] = 'VARIACION DE FIN:'
    ws['C16'].font = header_font
    
    ws['A17'] = 'DURACION REAL DE EJECUCION (DIAS):'
    ws['A17'].font = header_font
    
    # Tabla de items
    ws['A19'] = 'ÍTEM'
    ws['A19'].font = header_font
    ws['A19'].alignment = center_alignment
    ws['A19'].border = thin_border
    
    ws['B19'] = 'CÓDIGO DE MUESTRA'
    ws['B19'].font = header_font
    ws['B19'].alignment = center_alignment
    ws['B19'].border = thin_border
    
    ws['C19'] = 'DESCRIPCIÓN'
    ws['C19'].font = header_font
    ws['C19'].alignment = center_alignment
    ws['C19'].border = thin_border
    
    ws['D19'] = 'CANTIDAD'
    ws['D19'].font = header_font
    ws['D19'].alignment = center_alignment
    ws['D19'].border = thin_border
    
    # Footer
    ws['A21'] = 'OBSERVACIONES:'
    ws['A21'].font = header_font
    
    ws['A22'] = 'O/T APERTURADA POR:'
    ws['A22'].font = header_font
    ws['C22'] = 'OT DESIGADA A:'
    ws['C22'].font = header_font
    
    # Ajustar anchos de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    # Aplicar bordes a las celdas principales
    for row in range(19, 20):  # Header de la tabla
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    
    # Guardar template
    wb.save('backend/templates/orden_trabajo_template.xlsx')
    print("Template de orden de trabajo creado exitosamente")

if __name__ == "__main__":
    crear_template_orden_trabajo()